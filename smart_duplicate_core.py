from __future__ import annotations

import hashlib
import io
import json
import math
import re
import sqlite3
import threading
import time
import tempfile
import urllib.error
import urllib.request
import zipfile
from collections import Counter, defaultdict
from contextlib import contextmanager
from dataclasses import dataclass
from datetime import datetime, timezone
from difflib import SequenceMatcher
from pathlib import Path
from typing import Any

try:
    from langdetect import DetectorFactory, LangDetectException, detect
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    DetectorFactory = None
    LangDetectException = Exception
    detect = None

try:
    from rapidfuzz import fuzz
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    fuzz = None

try:
    import nltk
    from nltk.tokenize import sent_tokenize as nltk_sent_tokenize
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    nltk = None
    nltk_sent_tokenize = None

try:
    import openpyxl
    from openpyxl.styles import PatternFill
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    openpyxl = None
    PatternFill = None

try:
    import spacy
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    spacy = None

try:
    from sentence_transformers import SentenceTransformer
except ModuleNotFoundError:  # pragma: no cover - optional dependency
    SentenceTransformer = None


DetectorFactory = DetectorFactory
if DetectorFactory is not None:
    DetectorFactory.seed = 0

ROOT = Path(__file__).resolve().parent
DATA_DIR = ROOT / "data"
DB_PATH = DATA_DIR / "smart_duplicate.sqlite3"
MAX_WORDS = 50_000
MIN_WORDS_WARN = 20
MIN_WORDS_REJECT = 1
DOC_ID_RE = re.compile(r"/document/d/([a-zA-Z0-9_-]{25,50})")
URL_RE = re.compile(r"https?://\S+")
HTML_TAG_RE = re.compile(r"<[^>]+>")
ZW_RE = re.compile(r"[\u200B-\u200D\uFEFF]")
CONTROL_RE = re.compile(r"[\x00-\x08\x0B-\x1F\x7F]")
WORD_RE = re.compile(r"[A-Za-z0-9]+(?:[-'][A-Za-z0-9]+)?")
SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+|\n+")
SECTION_HINTS = {
    "intro": ["intro", "introduction", "overview", "about", "description"],
    "overview": ["overview", "about", "introduction", "description"],
    "features": ["features", "benefits", "key features", "advantages"],
    "benefits": ["benefits", "advantages", "why choose"],
    "specifications": ["specifications", "technical specs", "product details", "dimensions", "materials", "specs"],
    "use_cases": ["use cases", "applications", "when to use", "ideal for", "applications", "scenarios"],
    "faq": ["faq", "faqs", "frequently asked questions", "common questions"],
    "conclusion": ["conclusion", "summary", "final thoughts", "closing"],
    "description": ["description", "product description", "overview"],
    "unknown_section": ["unknown"],
}
SECTION_ALIASES = {
    "intro": "intro",
    "introduction": "intro",
    "overview": "intro",
    "about": "intro",
    "description": "intro",
    "features": "features",
    "benefits": "features",
    "applications": "use_cases",
    "use cases": "use_cases",
    "when to use": "use_cases",
    "ideal for": "use_cases",
    "faq": "faq",
    "faqs": "faq",
    "frequently asked questions": "faq",
    "common questions": "faq",
    "specifications": "specifications",
    "tech specs": "specifications",
    "product details": "specifications",
    "dimensions": "specifications",
    "materials": "specifications",
    "conclusion": "conclusion",
    "summary": "conclusion",
    "final thoughts": "conclusion",
}
FULLY_ALLOWED_SECTIONS = {"specifications", "product details", "technical specs", "dimensions", "materials", "title", "product name"}
DEFAULT_THRESHOLDS = {
    "intro": 45.0,
    "description": 45.0,
    "use_cases": 50.0,
    "applications": 50.0,
    "benefits": 50.0,
    "features": 50.0,
    "faq": 40.0,
    "conclusion": 45.0,
    "unknown_section": 48.0,
    "phần_đầu": 48.0,
    "phần_giữa": 48.0,
    "phần_cuối": 48.0,
}
SPEC_PATTERNS = [
    ("measurements", r"\b\d+(\.\d+)?\s*(ml|oz|l|cl|g|kg|cm|mm|m|inch|ft)\b"),
    ("quantities", r"\b\d+\s*(pcs|pieces|pack|units|case|box|carton)\b"),
    ("dimensions", r"\b\d+\s*[xX×]\s*\d+(\s*[xX×]\s*\d+)?\b"),
    ("prices", r"[£$€¥]\s*\d+(\.\d{2})?|\d+(\.\d{2})?\s*[pP]"),
    ("percentages", r"\b\d+(\.\d+)?%\b"),
    ("codes", r"[A-Z]{2,}-\d{3,}|\b[A-Z]{3,}\d{3,}\b"),
]
GENERIC_OPENERS = ("this product is", "the x provides", "designed for", "this container is", "the product offers")
GENERIC_CLOSERS = ("delivery service", "professional use", "food service", "foodservice use", "takeaway service")
INDUSTRY_TERMS = {
    "takeaway",
    "delivery",
    "food service",
    "foodservice",
    "professional use",
    "portion control",
    "lid closure",
    "heat retention",
}
SEMANTIC_ALIASES = [
    (r"\bcourier transport\b", "delivery"),
    (r"\btransport\b", "delivery"),
    (r"\btransit\b", "delivery"),
    (r"\beasy to stack\b", "stackable"),
    (r"\bstays stackable\b", "stackable"),
    (r"\blid fit\b", "lid closure"),
    (r"\bsecure lid fit\b", "lid closure"),
    (r"\bsecure closure\b", "lid closure"),
    (r"\bsuitable for\b", "ideal for"),
    (r"\bsupports\b", "helps"),
    (r"\bpack\b", "container"),
    (r"\bitem\b", "product"),
    (r"\bworkflow(s)?\b", "operations"),
]
STOPWORDS = {
    "the",
    "a",
    "an",
    "is",
    "are",
    "for",
    "and",
    "or",
    "to",
    "of",
    "with",
    "in",
    "on",
    "that",
    "this",
    "these",
    "those",
    "it",
    "its",
    "be",
    "as",
    "by",
    "from",
    "at",
    "into",
    "their",
    "they",
    "will",
    "can",
    "may",
}
COMMON_WORDS = STOPWORDS | {
    "strong",
    "durable",
    "useful",
    "good",
    "great",
    "ideal",
    "suitable",
    "high",
    "low",
    "compact",
    "practical",
    "safe",
    "easy",
    "simple",
}
RESULT_LEVELS = {"conflict": "🔴", "review": "🟡", "ok": "✅", "skip": "✅", "allowed": "✅"}


def utc_now() -> str:
    return datetime.now(timezone.utc).isoformat()


def ensure_data_dir() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)


def normalize_space(text: str) -> str:
    return re.sub(r"\s+", " ", text or "").strip()


def normalize_compare(text: str) -> str:
    lowered = normalize_space(text).lower()
    lowered = re.sub(r"[^\w\s%£$€¥.-]", " ", lowered)
    return re.sub(r"\s+", " ", lowered).strip()


def sanitize_text(text: str) -> str:
    if text is None:
        return ""
    cleaned = text.replace("\r\n", "\n").replace("\r", "\n")
    cleaned = cleaned.replace("“", '"').replace("”", '"').replace("’", "'").replace("‘", "'")
    cleaned = re.sub(r"<script.*?</script>", " ", cleaned, flags=re.IGNORECASE | re.DOTALL)
    cleaned = HTML_TAG_RE.sub(" ", cleaned)
    cleaned = ZW_RE.sub("", cleaned)
    cleaned = CONTROL_RE.sub(" ", cleaned)
    return re.sub(r"\n{3,}", "\n\n", cleaned).strip()


def tokenize(text: str) -> list[str]:
    return [token.lower() for token in WORD_RE.findall(text or "")]


def meaningful_tokens(text: str) -> list[str]:
    return [token for token in tokenize(text) if token not in STOPWORDS and token != "spec"]


def semantic_normalize(text: str) -> str:
    normalized = normalize_compare(text)
    for pattern, replacement in SEMANTIC_ALIASES:
        normalized = re.sub(pattern, replacement, normalized, flags=re.IGNORECASE)
    return normalize_space(normalized)


def semantic_tokens(text: str) -> list[str]:
    return [token for token in tokenize(semantic_normalize(text)) if token not in STOPWORDS and token != "spec"]


def text_hash(text: str) -> str:
    return hashlib.sha256(normalize_space(text).encode("utf-8")).hexdigest()


def sent_tokenize_safe(text: str) -> list[str]:
    normalized = sanitize_text(text)
    if not normalized:
        return []
    if nltk_sent_tokenize is not None:
        try:
            return [normalize_space(item) for item in nltk_sent_tokenize(normalized) if normalize_space(item)]
        except LookupError:  # pragma: no cover - depends on punkt download
            pass
    return [normalize_space(item) for item in SENTENCE_SPLIT_RE.split(normalized) if normalize_space(item)]


def detect_language_safe(text: str) -> str:
    sample = " ".join(tokenize(text)[:300])
    if len(sample.split()) < 8 or detect is None:
        return "unknown"
    try:
        return detect(sample)
    except LangDetectException:  # pragma: no cover - depends on input
        return "unknown"


def fuzzy_ratio(left: str, right: str) -> float:
    if not left or not right:
        return 0.0
    if fuzz is not None:
        return float(fuzz.ratio(left, right))
    return SequenceMatcher(None, normalize_compare(left), normalize_compare(right)).ratio() * 100.0


def cosine_similarity_from_counters(left: Counter[str], right: Counter[str]) -> float:
    if not left or not right:
        return 0.0
    shared = set(left) & set(right)
    numerator = sum(left[token] * right[token] for token in shared)
    norm_left = math.sqrt(sum(value * value for value in left.values()))
    norm_right = math.sqrt(sum(value * value for value in right.values()))
    if not norm_left or not norm_right:
        return 0.0
    return numerator / (norm_left * norm_right)


def word_trigrams(tokens: list[str]) -> set[tuple[str, str, str]]:
    if len(tokens) < 3:
        return set()
    return {tuple(tokens[index : index + 3]) for index in range(0, len(tokens) - 2)}


def slugify(value: str) -> str:
    return re.sub(r"[^a-z0-9]+", "-", normalize_compare(value)).strip("-") or "untitled"


def normalize_allowed_phrase(phrase: str) -> str:
    cleaned = normalize_space(phrase).strip(" .,:;!?-–—()[]{}\"'")
    tokens = tokenize(cleaned)
    if not tokens:
        return ""
    if len(tokens) > 10:
        return ""
    if all(token in STOPWORDS for token in tokens):
        return ""
    if len(tokens) == 1:
        token = tokens[0]
        if len(token) < 4:
            return ""
        if token in STOPWORDS or token in {"product", "title", "intro", "overview", "features", "benefits", "faq"}:
            return ""
    return cleaned


def compile_phrase_pattern(phrase: str) -> re.Pattern[str]:
    escaped = re.escape(phrase)
    if phrase and phrase[0].isalnum() and phrase[-1].isalnum():
        return re.compile(rf"(?<!\w){escaped}(?!\w)", re.IGNORECASE)
    return re.compile(escaped, re.IGNORECASE)


def google_doc_id_from_url(url: str) -> str:
    match = DOC_ID_RE.search(url or "")
    if match:
        return match.group(1)
    return ""


def fetch_google_doc_text(url: str) -> tuple[str, list[str]]:
    warnings: list[str] = []
    doc_id = google_doc_id_from_url(url)
    if not doc_id:
        raise ValueError("URL Google Docs khong hop le")
    export_url = f"https://docs.google.com/document/d/{doc_id}/export?format=txt"
    retries = 0
    while retries < 3:
        request = urllib.request.Request(export_url, headers={"User-Agent": "smart-duplicate-checker/1.0"})
        try:
            with urllib.request.urlopen(request, timeout=10) as response:
                payload = response.read().decode("utf-8", errors="replace")
            return payload, warnings
        except urllib.error.HTTPError as exc:
            if exc.code == 403:
                raise ValueError("Doc nay can share 'Anyone with link can view'") from exc
            if exc.code == 404:
                raise ValueError("Khong tim thay doc. Kiem tra lai URL") from exc
            if exc.code == 429 and retries == 0:
                time.sleep(3)
                retries += 1
                continue
            raise ValueError(f"Khong tai duoc doc: HTTP {exc.code}") from exc
        except urllib.error.URLError as exc:
            retries += 1
            if retries < 3:
                warnings.append("Dang thu tai lai Google Docs...")
                time.sleep(1 + retries)
                continue
            raise ValueError("Khong tai duoc doc. Vui long thu lai hoac dan raw text") from exc
    raise ValueError("Khong tai duoc doc. Vui long thu lai hoac dan raw text")


def first_non_empty_line(text: str) -> str:
    for line in sanitize_text(text).splitlines():
        line = normalize_space(line)
        if line:
            return line
    return ""


def canonical_section_name(raw_name: str, template_sections: list[str] | None = None) -> str:
    candidate = normalize_space(raw_name).lower().strip(":")
    if not candidate:
        return "unknown_section"
    if candidate in SECTION_ALIASES:
        return SECTION_ALIASES[candidate]
    for canonical, aliases in SECTION_HINTS.items():
        for alias in aliases:
            if alias in candidate:
                return canonical
    best_name = "unknown_section"
    best_score = 0.0
    options = set(SECTION_ALIASES) | set(SECTION_HINTS)
    if template_sections:
        options.update(item.lower() for item in template_sections)
    for option in options:
        score = fuzzy_ratio(candidate, option.lower())
        if score > best_score:
            best_score = score
            best_name = SECTION_ALIASES.get(option.lower(), option.lower())
    return best_name if best_score >= 75.0 else "unknown_section"


def is_heading_candidate(lines: list[str], index: int, template_sections: list[str] | None = None) -> bool:
    line = lines[index].strip()
    if not line:
        return False
    if line.startswith("#"):
        return True
    if len(line) < 60 and index + 1 < len(lines) and not lines[index + 1].strip():
        return True
    token_count = len(line.split())
    if token_count <= 12 and line.upper() == line and any(char.isalpha() for char in line):
        return True
    canonical = canonical_section_name(line, template_sections)
    return canonical != "unknown_section"


def parse_sections(text: str, template_sections: list[str] | None = None) -> tuple[str, dict[str, dict[str, Any]], list[str]]:
    cleaned = sanitize_text(text)
    warnings: list[str] = []
    if not cleaned:
        raise ValueError("Bai nay trong, vui long kiem tra lai")
    lines = cleaned.splitlines()
    title = first_non_empty_line(cleaned)[:160] or "Khong co tieu de"
    sections: dict[str, dict[str, Any]] = {}
    current_heading = "intro"
    buffer: list[str] = []
    found_heading = False

    def flush_buffer() -> None:
        nonlocal buffer, current_heading
        content = sanitize_text("\n".join(buffer))
        if not content:
            buffer = []
            return
        canonical = canonical_section_name(current_heading, template_sections)
        section_name = canonical if canonical != "unknown_section" else current_heading or "unknown_section"
        existing = sections.get(section_name, {"title": current_heading, "text": ""})
        joined = "\n\n".join(part for part in [existing.get("text", ""), content] if part)
        sections[section_name] = {"title": current_heading, "text": joined}
        buffer = []

    for index, raw_line in enumerate(lines):
        line = raw_line.strip()
        if is_heading_candidate(lines, index, template_sections):
            found_heading = True
            flush_buffer()
            current_heading = line.lstrip("#").strip().strip(":")
            continue
        buffer.append(raw_line)
    flush_buffer()

    if not found_heading or not sections:
        paragraphs = [normalize_space(chunk) for chunk in re.split(r"\n\s*\n", cleaned) if normalize_space(chunk)]
        if not paragraphs:
            paragraphs = [normalize_space(cleaned)]
        groups = ["phần_đầu", "phần_giữa", "phần_cuối"]
        chunk_size = max(1, math.ceil(len(paragraphs) / 3))
        sections = {}
        for index, group_name in enumerate(groups):
            part = paragraphs[index * chunk_size : (index + 1) * chunk_size]
            if part:
                sections[group_name] = {"title": group_name, "text": "\n\n".join(part)}
        warnings.append("Khong phat hien duoc heading — dang so sanh theo doan van")

    normalized_sections = {}
    for name, payload in sections.items():
        canonical = canonical_section_name(name, template_sections)
        final_name = canonical if canonical != "unknown_section" else name
        normalized_sections[final_name] = {
            "title": payload.get("title", name),
            "text": sanitize_text(payload.get("text", "")),
            "canonical_name": final_name,
        }
    return title, normalized_sections, warnings


def extract_pattern_examples(text: str) -> list[dict[str, Any]]:
    cleaned = sanitize_text(text)
    results = []
    for label, pattern in SPEC_PATTERNS:
        matches = []
        for match in re.finditer(pattern, cleaned, flags=re.IGNORECASE):
            snippet = normalize_space(match.group(0))
            if snippet and snippet not in matches:
                matches.append(snippet)
            if len(matches) >= 5:
                break
        results.append({"label": label, "regex": pattern, "examples": matches})
    return results


def extract_brand_terms(text: str, title: str) -> list[str]:
    phrases: list[str] = []
    cleaned = sanitize_text(text)
    title_line = normalize_space(title)
    if not title_line:
        return phrases
    title_tokens = set(tokenize(title_line))
    phrase_candidates: set[str] = set()
    if spacy is not None:
        try:
            nlp = spacy.load("en_core_web_sm")
            doc = nlp(cleaned[:5000])
            for chunk in doc.noun_chunks:
                candidate = normalize_space(chunk.text)
                if candidate and any(word[0].isupper() for word in candidate.split() if word):
                    phrase_candidates.add(candidate)
        except Exception:  # pragma: no cover - model may be missing
            pass
    if not phrase_candidates:
        regex = re.compile(r"\b(?:[A-Z][A-Za-z0-9/-]+(?:\s+[A-Z][A-Za-z0-9/-]+){0,3})\b")
        phrase_candidates.update(match.group(0) for match in regex.finditer(cleaned))
    lowered_text = cleaned.lower()
    for candidate in sorted(phrase_candidates):
        normalized_candidate = normalize_allowed_phrase(candidate)
        if not normalized_candidate:
            continue
        candidate_tokens = tokenize(normalized_candidate)
        if not candidate_tokens:
            continue
        if not set(candidate_tokens).issubset(title_tokens):
            continue
        match_count = len(compile_phrase_pattern(normalized_candidate).findall(cleaned))
        if len(candidate_tokens) == 1 and candidate_tokens[0] in {"kraft", "paper", "container", "bowl", "lid"}:
            continue
        if match_count >= 3:
            phrases.append(normalized_candidate)
    return phrases[:20]


def detect_template_phrases(text: str) -> list[str]:
    phrases: list[str] = []
    for sentence in sent_tokenize_safe(text):
        lowered = sentence.lower()
        tokens = tokenize(sentence)
        if not tokens:
            continue
        if re.search(r"[\[\{].*[\]\}]", sentence):
            phrases.append(sentence)
            continue
        content_words = [token for token in tokens if token not in STOPWORDS]
        common_ratio = 0.0
        if content_words:
            common_ratio = sum(1 for token in content_words if token in COMMON_WORDS) / len(content_words)
        if 5 <= len(tokens) < 20 and common_ratio >= 0.7:
            normalized_sentence = normalize_allowed_phrase(sentence)
            if normalized_sentence:
                phrases.append(normalized_sentence)
            continue
        for opener in GENERIC_OPENERS:
            if lowered.startswith(opener):
                phrases.append(opener)
        for closer in GENERIC_CLOSERS:
            if lowered.endswith(closer):
                phrases.append(closer)
    deduped = []
    seen = set()
    for phrase in phrases:
        normalized_phrase = normalize_allowed_phrase(phrase)
        lowered = normalize_compare(normalized_phrase)
        if lowered and lowered not in seen:
            seen.add(lowered)
            deduped.append(normalized_phrase)
    return deduped[:40]


def detect_fully_allowed_sections(sections: dict[str, dict[str, Any]]) -> list[str]:
    fully_allowed: list[str] = []
    spec_patterns = [pattern for _label, pattern in SPEC_PATTERNS]
    for name, payload in sections.items():
        canonical = canonical_section_name(name)
        text = payload.get("text", "")
        if canonical in {"specifications"} or name.lower() in FULLY_ALLOWED_SECTIONS:
            fully_allowed.append(canonical)
            continue
        cleaned = re.sub(r"\s+", " ", text)
        non_pattern = cleaned
        for pattern in spec_patterns:
            non_pattern = re.sub(pattern, " ", non_pattern, flags=re.IGNORECASE)
        non_pattern = re.sub(r"[\W_]+", " ", non_pattern)
        if len([token for token in tokenize(non_pattern) if token not in STOPWORDS]) == 0:
            fully_allowed.append(canonical)
    return sorted(set(item for item in fully_allowed if item))


def default_allowed_zone_config(template_text: str, sections: dict[str, dict[str, Any]], title: str) -> dict[str, Any]:
    pattern_entries = extract_pattern_examples(template_text)
    brand_terms = extract_brand_terms(template_text, title)
    template_phrases = detect_template_phrases(template_text)
    allowed_phrases = []
    seen = set()
    for candidate in brand_terms + template_phrases:
        normalized_candidate = normalize_allowed_phrase(candidate)
        lowered_candidate = normalize_compare(normalized_candidate)
        if not lowered_candidate or lowered_candidate in seen:
            continue
        seen.add(lowered_candidate)
        allowed_phrases.append(normalized_candidate)
    fully_allowed_sections = detect_fully_allowed_sections(sections)
    return {
        "allowed_patterns_regex": pattern_entries,
        "allowed_phrases": allowed_phrases,
        "fully_allowed_sections": fully_allowed_sections,
        "section_thresholds": dict(DEFAULT_THRESHOLDS),
        "non_paraphraseable_terms": sorted(INDUSTRY_TERMS),
        "template_sections": [payload.get("canonical_name", name) for name, payload in sections.items()],
    }


def strip_allowed_content(section_name: str, text: str, config: dict[str, Any]) -> dict[str, Any]:
    section_canonical = canonical_section_name(section_name)
    if section_canonical in set(config.get("fully_allowed_sections", [])):
        return {
            "protected_text": "",
            "strip_log": {"removed_sentences": 0, "removed_phrases": [], "removed_patterns": 0},
            "status": "allowed",
            "note": "Allowed zone - khong check",
            "meaningful_word_count": 0,
        }
    working = sanitize_text(text)
    removed_phrases: list[str] = []
    removed_patterns = 0
    allowed_phrases = []
    for phrase in config.get("allowed_phrases", []):
        normalized_phrase = normalize_allowed_phrase(phrase)
        if not normalized_phrase:
            continue
        allowed_phrases.append(normalized_phrase)
    allowed_phrases = sorted(dict.fromkeys(allowed_phrases), key=len, reverse=True)
    for phrase in allowed_phrases:
        pattern = compile_phrase_pattern(phrase)
        if pattern.search(working):
            removed_phrases.append(phrase)
            working = pattern.sub(" ", working)
    for entry in config.get("allowed_patterns_regex", []):
        regex = entry.get("regex", "") if isinstance(entry, dict) else str(entry)
        if not regex:
            continue
        matches = re.findall(regex, working, flags=re.IGNORECASE)
        removed_patterns += len(matches)
        working = re.sub(regex, " [SPEC] ", working, flags=re.IGNORECASE)
    removed_sentences = 0
    kept_sentences: list[str] = []
    for sentence in sent_tokenize_safe(working):
        original_tokens = sentence.split()
        if not original_tokens:
            continue
        temp_sentence = sentence
        for phrase in allowed_phrases:
            temp_sentence = compile_phrase_pattern(phrase).sub(" ", temp_sentence)
        for entry in config.get("allowed_patterns_regex", []):
            regex = entry.get("regex", "") if isinstance(entry, dict) else str(entry)
            temp_sentence = re.sub(regex, " ", temp_sentence, flags=re.IGNORECASE)
        remaining = meaningful_tokens(temp_sentence)
        if len(remaining) / max(len(original_tokens), 1) < 0.35:
            removed_sentences += 1
            continue
        if len(remaining) < 8:
            removed_sentences += 1
            continue
        kept_sentences.append(normalize_space(sentence))
    protected_text = " ".join(kept_sentences).strip()
    meaningful_count = len(meaningful_tokens(protected_text))
    if meaningful_count < 12:
        return {
            "protected_text": protected_text,
            "strip_log": {
                "removed_sentences": removed_sentences,
                "removed_phrases": removed_phrases[:20],
                "removed_patterns": removed_patterns,
            },
            "status": "skip",
            "note": "Bo qua - qua ngan",
            "meaningful_word_count": meaningful_count,
        }
    return {
        "protected_text": protected_text,
        "strip_log": {
            "removed_sentences": removed_sentences,
            "removed_phrases": removed_phrases[:20],
            "removed_patterns": removed_patterns,
        },
        "status": "ok",
        "note": f"Da bo qua {removed_sentences} cau chua allowed content",
        "meaningful_word_count": meaningful_count,
    }


def tf_counter(text: str) -> Counter[str]:
    return Counter(meaningful_tokens(text))


def longest_common_sentence_score(left_sentences: list[str], right_sentences: list[str]) -> tuple[float, list[str]]:
    if not left_sentences or not right_sentences:
        return 0.0, []
    normalized_left = [normalize_compare(item) for item in left_sentences]
    normalized_right = [normalize_compare(item) for item in right_sentences]
    dp = [[0] * (len(normalized_right) + 1) for _ in range(len(normalized_left) + 1)]
    lcs_sentences: list[str] = []
    for i in range(1, len(normalized_left) + 1):
        for j in range(1, len(normalized_right) + 1):
            if normalized_left[i - 1] == normalized_right[j - 1]:
                dp[i][j] = dp[i - 1][j - 1] + 1
            else:
                dp[i][j] = max(dp[i - 1][j], dp[i][j - 1])
    i, j = len(normalized_left), len(normalized_right)
    while i > 0 and j > 0:
        if normalized_left[i - 1] == normalized_right[j - 1]:
            lcs_sentences.append(left_sentences[i - 1])
            i -= 1
            j -= 1
        elif dp[i - 1][j] >= dp[i][j - 1]:
            i -= 1
        else:
            j -= 1
    lcs_sentences.reverse()
    score = (len(lcs_sentences) / max(1, min(len(left_sentences), len(right_sentences)))) * 100.0
    return score, lcs_sentences


def best_matching_sentence(left_sentences: list[str], right_sentences: list[str]) -> tuple[str, str]:
    best_left = left_sentences[0] if left_sentences else ""
    best_right = right_sentences[0] if right_sentences else ""
    best_score = 0.0
    for left in left_sentences:
        left_norm = normalize_compare(left)
        for right in right_sentences:
            score = SequenceMatcher(None, left_norm, normalize_compare(right)).ratio()
            if score > best_score:
                best_score = score
                best_left = left
                best_right = right
    return best_left[:280], best_right[:280]


class SemanticModel:
    def __init__(self) -> None:
        self._models: dict[str, Any] = {}
        self._lock = threading.Lock()

    def available(self) -> bool:
        return SentenceTransformer is not None

    def _load(self, multilingual: bool) -> Any | None:
        if SentenceTransformer is None:
            return None
        key = "multi" if multilingual else "en"
        with self._lock:
            if key in self._models:
                return self._models[key]
            try:
                model_name = (
                    "paraphrase-multilingual-MiniLM-L12-v2" if multilingual else "paraphrase-MiniLM-L6-v2"
                )
                self._models[key] = SentenceTransformer(model_name)
            except Exception:  # pragma: no cover - depends on model availability
                self._models[key] = None
            return self._models[key]

    def similarity(self, left: str, right: str, *, multilingual: bool) -> float | None:
        model = self._load(multilingual)
        if model is None:
            return None
        try:
            embeddings = model.encode([left, right], normalize_embeddings=True)
            return float(sum(embeddings[0][index] * embeddings[1][index] for index in range(len(embeddings[0])))) * 100.0
        except Exception:  # pragma: no cover - depends on runtime model state
            return None


@dataclass
class ComparisonOutcome:
    section: str
    risk_score: float
    level: str
    flagged_sentences: dict[str, Any]


class SmartDuplicateService:
    def __init__(self, db_path: Path | None = None):
        ensure_data_dir()
        self.db_path = Path(db_path or DB_PATH)
        self.semantic_model = SemanticModel()
        self.project_locks: dict[int, threading.Lock] = defaultdict(threading.Lock)
        self._init_db()
        self.seed_example_project()

    @contextmanager
    def connection(self):
        conn = sqlite3.connect(self.db_path)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA journal_mode=WAL;")
        conn.execute("PRAGMA foreign_keys=ON;")
        try:
            yield conn
            conn.commit()
        except Exception:
            conn.rollback()
            raise
        finally:
            conn.close()

    def _init_db(self) -> None:
        with self.connection() as conn:
            conn.execute("PRAGMA integrity_check;")
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS projects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    template_doc_text TEXT NOT NULL DEFAULT '',
                    allowed_zone_config_json TEXT NOT NULL DEFAULT '{}',
                    updated_at TEXT NOT NULL DEFAULT ''
                );
                CREATE TABLE IF NOT EXISTS documents (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    title TEXT NOT NULL,
                    raw_text TEXT NOT NULL,
                    sections_json TEXT NOT NULL,
                    added_at TEXT NOT NULL,
                    content_hash TEXT NOT NULL DEFAULT '',
                    source_url TEXT NOT NULL DEFAULT '',
                    doc_role TEXT NOT NULL DEFAULT 'check',
                    approval_status TEXT NOT NULL DEFAULT 'approved',
                    warnings_json TEXT NOT NULL DEFAULT '[]',
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS comparisons (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    doc_a_id INTEGER NOT NULL,
                    doc_b_id INTEGER NOT NULL,
                    section TEXT NOT NULL,
                    risk_score REAL NOT NULL,
                    flagged_sentences_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    level TEXT NOT NULL DEFAULT 'ok',
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                    FOREIGN KEY(doc_a_id) REFERENCES documents(id) ON DELETE CASCADE,
                    FOREIGN KEY(doc_b_id) REFERENCES documents(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS embeddings_cache (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    document_id INTEGER NOT NULL,
                    section TEXT NOT NULL,
                    model_name TEXT NOT NULL,
                    text_hash TEXT NOT NULL,
                    embedding_json TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    UNIQUE(document_id, section, model_name)
                );
                CREATE TABLE IF NOT EXISTS project_jobs (
                    project_id INTEGER PRIMARY KEY,
                    status TEXT NOT NULL,
                    progress_current INTEGER NOT NULL DEFAULT 0,
                    progress_total INTEGER NOT NULL DEFAULT 0,
                    message TEXT NOT NULL DEFAULT '',
                    updated_at TEXT NOT NULL
                );
                """
            )
            document_columns = {
                row["name"] for row in conn.execute("PRAGMA table_info(documents)").fetchall()
            }
            if "doc_role" not in document_columns:
                conn.execute("ALTER TABLE documents ADD COLUMN doc_role TEXT NOT NULL DEFAULT 'check'")
            if "approval_status" not in document_columns:
                conn.execute("ALTER TABLE documents ADD COLUMN approval_status TEXT NOT NULL DEFAULT 'approved'")
            conn.execute(
                """
                UPDATE documents
                SET approval_status = CASE
                    WHEN approval_status IS NULL OR approval_status = '' THEN 'approved'
                    ELSE approval_status
                END
                """
            )

    def seed_example_project(self) -> None:
        with self.connection() as conn:
            total = conn.execute("SELECT COUNT(*) AS total FROM projects").fetchone()["total"]
            if int(total) > 0:
                return
        template_text = """
        KRAFT ROUND BOWL 900ML

        OVERVIEW
        This product is designed for professional takeaway and delivery use. The kraft bowl provides reliable food service performance.

        FEATURES
        The bowl has a secure lid fit and helps reduce leaks during delivery. It stays stackable in busy kitchens.

        USE CASES
        Ideal for salad, noodle, and rice bowl service in takeaway workflows.

        FAQ
        Can it handle delivery? Yes. The secure closure helps keep contents protected in transit.

        CONCLUSION
        Designed for professional use and dependable takeaway service.
        """
        project = self.create_project("Vi du")
        self.set_template_from_text(project["id"], template_text)
        self.add_document_from_text(
            project["id"],
            "Vi du 1",
            """
            Kraft Round Bowl 950ml

            Overview
            This product is designed for professional takeaway and delivery use. The kraft bowl provides reliable food service performance.

            Features
            The bowl has a secure lid fit and helps reduce leaks during delivery. It stays stackable in busy kitchens.
            """,
            source_url="",
            source_only=True,
        )
        self.add_document_from_text(
            project["id"],
            "Vi du 2",
            """
            Kraft Round Bowl 1200ml

            Overview
            Suitable for larger takeaway meals and deli display operations.

            Features
            A firm lid fit helps protect food during courier transport while the bowl remains easy to stack.
            """,
            source_url="",
        )

    def create_project(self, name: str) -> dict[str, Any]:
        now = utc_now()
        with self.connection() as conn:
            cursor = conn.execute(
                """
                INSERT INTO projects (name, created_at, template_doc_text, allowed_zone_config_json, updated_at)
                VALUES (?, ?, '', '{}', ?)
                """,
                (normalize_space(name) or "Project moi", now, now),
            )
            project_id = int(cursor.lastrowid)
        return self.get_project(project_id)

    def get_project(self, project_id: int) -> dict[str, Any]:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay project")
            return self._project_payload(conn, row)

    def _project_payload(self, conn: sqlite3.Connection, row: sqlite3.Row) -> dict[str, Any]:
        doc_count = conn.execute("SELECT COUNT(*) AS total FROM documents WHERE project_id = ?", (row["id"],)).fetchone()["total"]
        source_doc_count = conn.execute(
            "SELECT COUNT(*) AS total FROM documents WHERE project_id = ? AND doc_role = 'source'",
            (row["id"],),
        ).fetchone()["total"]
        pending_doc_count = conn.execute(
            "SELECT COUNT(*) AS total FROM documents WHERE project_id = ? AND approval_status = 'pending'",
            (row["id"],),
        ).fetchone()["total"]
        conflict_count = conn.execute(
            "SELECT COUNT(*) AS total FROM comparisons WHERE project_id = ? AND level = 'conflict'",
            (row["id"],),
        ).fetchone()["total"]
        template_title = first_non_empty_line(row["template_doc_text"]) if row["template_doc_text"] else ""
        updated_at = row["updated_at"] or row["created_at"]
        return {
            "id": int(row["id"]),
            "name": row["name"],
            "created_at": row["created_at"],
            "updated_at": updated_at,
            "template_name": template_title or "Chua co template",
            "doc_count": int(doc_count),
            "source_doc_count": int(source_doc_count),
            "pending_doc_count": int(pending_doc_count),
            "conflict_count": int(conflict_count),
            "allowed_zone_config": json.loads(row["allowed_zone_config_json"] or "{}"),
            "has_template": bool(row["template_doc_text"]),
        }

    def list_projects(self) -> list[dict[str, Any]]:
        with self.connection() as conn:
            rows = conn.execute("SELECT * FROM projects ORDER BY updated_at DESC, created_at DESC").fetchall()
            return [self._project_payload(conn, row) for row in rows]

    def rename_project(self, project_id: int, name: str) -> dict[str, Any]:
        with self.connection() as conn:
            conn.execute("UPDATE projects SET name = ?, updated_at = ? WHERE id = ?", (normalize_space(name), utc_now(), project_id))
        return self.get_project(project_id)

    def duplicate_project(self, project_id: int) -> dict[str, Any]:
        project = self.get_project(project_id)
        new_project = self.create_project(f"{project['name']} - ban sao")
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            conn.execute(
                "UPDATE projects SET template_doc_text = ?, allowed_zone_config_json = ?, updated_at = ? WHERE id = ?",
                (row["template_doc_text"], row["allowed_zone_config_json"], utc_now(), new_project["id"]),
            )
            docs = conn.execute("SELECT * FROM documents WHERE project_id = ? ORDER BY added_at ASC", (project_id,)).fetchall()
            id_map: dict[int, int] = {}
            for doc in docs:
                cursor = conn.execute(
                    """
                    INSERT INTO documents (
                        project_id, title, raw_text, sections_json, added_at,
                        content_hash, source_url, doc_role, approval_status, warnings_json
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_project["id"],
                        doc["title"],
                        doc["raw_text"],
                        doc["sections_json"],
                        doc["added_at"],
                        doc["content_hash"],
                        doc["source_url"],
                        doc["doc_role"] if "doc_role" in doc.keys() else "check",
                        doc["approval_status"] if "approval_status" in doc.keys() else "approved",
                        doc["warnings_json"],
                    ),
                )
                id_map[int(doc["id"])] = int(cursor.lastrowid)
            comparisons = conn.execute("SELECT * FROM comparisons WHERE project_id = ?", (project_id,)).fetchall()
            for comp in comparisons:
                conn.execute(
                    """
                    INSERT INTO comparisons (project_id, doc_a_id, doc_b_id, section, risk_score, flagged_sentences_json, created_at, level)
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_project["id"],
                        id_map.get(int(comp["doc_a_id"]), 0),
                        id_map.get(int(comp["doc_b_id"]), 0),
                        comp["section"],
                        comp["risk_score"],
                        comp["flagged_sentences_json"],
                        comp["created_at"],
                        comp["level"],
                    ),
                )
        return self.get_project(new_project["id"])

    def delete_project(self, project_id: int) -> None:
        with self.connection() as conn:
            conn.execute("DELETE FROM projects WHERE id = ?", (project_id,))

    def set_template_from_text(self, project_id: int, text: str, *, title: str = "") -> dict[str, Any]:
        clean_text = sanitize_text(text)
        if not meaningful_tokens(clean_text):
            raise ValueError("Bai nay trong, vui long kiem tra lai")
        template_title, sections, warnings = parse_sections(clean_text)
        config = default_allowed_zone_config(clean_text, sections, title or template_title)
        config["warnings"] = warnings
        with self.connection() as conn:
            conn.execute(
                "UPDATE projects SET template_doc_text = ?, allowed_zone_config_json = ?, updated_at = ? WHERE id = ?",
                (clean_text, json.dumps(config, ensure_ascii=False), utc_now(), project_id),
            )
        return {
            "template_title": title or template_title,
            "warnings": warnings,
            "allowed_zone": config,
        }

    def update_allowed_zone(self, project_id: int, config: dict[str, Any], *, recheck_all: bool = False) -> dict[str, Any]:
        normalized = {
            "allowed_patterns_regex": config.get("allowed_patterns_regex", []),
            "allowed_phrases": [normalize_space(item) for item in config.get("allowed_phrases", []) if normalize_space(item)],
            "fully_allowed_sections": [canonical_section_name(item) for item in config.get("fully_allowed_sections", []) if item],
            "section_thresholds": {**DEFAULT_THRESHOLDS, **config.get("section_thresholds", {})},
            "non_paraphraseable_terms": list(dict.fromkeys(config.get("non_paraphraseable_terms", sorted(INDUSTRY_TERMS)))),
            "template_sections": config.get("template_sections", []),
        }
        with self.connection() as conn:
            conn.execute(
                "UPDATE projects SET allowed_zone_config_json = ?, updated_at = ? WHERE id = ?",
                (json.dumps(normalized, ensure_ascii=False), utc_now(), project_id),
            )
        if recheck_all:
            self.recheck_project(project_id)
        return self.get_project(project_id)

    def _validate_doc_text(self, text: str) -> tuple[str, list[str]]:
        warnings: list[str] = []
        clean_text = sanitize_text(text)
        words = tokenize(clean_text)
        if len(words) <= MIN_WORDS_REJECT:
            raise ValueError("Bai nay trong, vui long kiem tra lai")
        if len(words) > MAX_WORDS:
            raise ValueError("Bai qua lon, toi da 50,000 tu")
        if len(words) < 50:
            warnings.append("Bai nay rat ngan, ket qua co the khong chinh xac")
        language = detect_language_safe(clean_text)
        if language not in {"en", "unknown"}:
            warnings.append("Phat hien ngon ngu khac, van tiep tuc check")
        if len(words) < MIN_WORDS_WARN:
            warnings.append("Bai nay rat ngan, ket qua co the khong chinh xac")
        table_like_lines = [line for line in clean_text.splitlines() if "|" in line or "\t" in line]
        if len(table_like_lines) >= 3 and len(words) < 120:
            warnings.append("Bai chu yeu la bang/hinh, text it")
        return clean_text, warnings

    def _project_config(self, conn: sqlite3.Connection, project_id: int) -> dict[str, Any]:
        row = conn.execute("SELECT allowed_zone_config_json FROM projects WHERE id = ?", (project_id,)).fetchone()
        if not row:
            raise ValueError("Khong tim thay project")
        config = json.loads(row["allowed_zone_config_json"] or "{}")
        if not config:
            config = {
                "allowed_patterns_regex": [],
                "allowed_phrases": [],
                "fully_allowed_sections": [],
                "section_thresholds": dict(DEFAULT_THRESHOLDS),
                "non_paraphraseable_terms": sorted(INDUSTRY_TERMS),
                "template_sections": [],
            }
        return config

    def _prepare_document(self, raw_text: str, config: dict[str, Any]) -> tuple[str, dict[str, Any], list[str]]:
        title, sections, warnings = parse_sections(raw_text, config.get("template_sections") or [])
        prepared_sections: dict[str, Any] = {}
        for name, payload in sections.items():
            prepared_sections[name] = {
                "title": payload["title"],
                "text": payload["text"],
                "canonical_name": payload.get("canonical_name", canonical_section_name(name)),
                "strip": strip_allowed_content(name, payload["text"], config),
            }
        return title, prepared_sections, warnings

    def _duplicate_document_message(self, row: sqlite3.Row) -> str:
        doc_role = row["doc_role"] if "doc_role" in row.keys() else "check"
        role_label = "nguồn đối chiếu" if doc_role == "source" else "bài đã check"
        doc_title = normalize_space(row["title"]) or f"ID {int(row['id'])}"
        added_at = row["added_at"] or "-"
        return (
            "Bài này đã có trong project. "
            f"Đã lưu dưới dạng {role_label}: \"{doc_title}\" (ID {int(row['id'])}, lúc {added_at}). "
            "Mở mục \"Kho bài đã import\" để tìm nhanh."
        )

    def add_document_from_text(
        self,
        project_id: int,
        title: str,
        text: str,
        *,
        source_url: str = "",
        source_only: bool = False,
    ) -> dict[str, Any]:
        with self.connection() as conn:
            config = self._project_config(conn, project_id)
            template_row = conn.execute("SELECT template_doc_text FROM projects WHERE id = ?", (project_id,)).fetchone()
            if not template_row or not template_row["template_doc_text"]:
                raise ValueError("Project nay chua co template. Vui long setup template truoc")
            clean_text, warnings = self._validate_doc_text(text)
            content_digest = text_hash(clean_text)
            existing_by_hash = conn.execute(
                "SELECT id, title, added_at, doc_role FROM documents WHERE project_id = ? AND content_hash = ? ORDER BY id DESC LIMIT 1",
                (project_id, content_digest),
            ).fetchone()
            if existing_by_hash:
                raise ValueError(self._duplicate_document_message(existing_by_hash))
            if source_url:
                existing_by_url = conn.execute(
                    """
                    SELECT id, title, added_at, doc_role
                    FROM documents
                    WHERE project_id = ? AND source_url = ? AND source_url != ''
                    ORDER BY id DESC
                    LIMIT 1
                    """,
                    (project_id, source_url),
                ).fetchone()
                if existing_by_url:
                    raise ValueError(self._duplicate_document_message(existing_by_url))
            parsed_title, sections, parse_warnings = self._prepare_document(clean_text, config)
            final_title = normalize_space(title) or parsed_title or f"Tai lieu {int(time.time())}"
            warnings.extend(parse_warnings)
            doc_role = "source" if source_only else "check"
            approval_status = "approved" if source_only else "pending"
            cursor = conn.execute(
                """
                INSERT INTO documents (
                    project_id, title, raw_text, sections_json, added_at,
                    content_hash, source_url, doc_role, approval_status, warnings_json
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    project_id,
                    final_title,
                    clean_text,
                    json.dumps(sections, ensure_ascii=False),
                    utc_now(),
                    content_digest,
                    source_url,
                    doc_role,
                    approval_status,
                    json.dumps(warnings, ensure_ascii=False),
                ),
            )
            document_id = int(cursor.lastrowid)
            conn.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (utc_now(), project_id))
            compare_target_count = conn.execute(
                "SELECT COUNT(*) AS total FROM documents WHERE project_id = ? AND doc_role = 'source' AND id != ?",
                (project_id, document_id),
            ).fetchone()["total"]
        if source_only:
            return {
                "document_id": document_id,
                "title": final_title,
                "queued": False,
                "warnings": warnings,
                "results": [],
                "message": "Đã import bài đã pass vào kho nguồn đối chiếu.",
            }
        if int(compare_target_count) > 50:
            self._set_job(project_id, "running", 0, int(compare_target_count), "Dang so sanh nen...")
            thread = threading.Thread(target=self._compare_document_background, args=(project_id, document_id), daemon=True)
            thread.start()
            return {
                "document_id": document_id,
                "title": final_title,
                "queued": True,
                "warnings": warnings,
                "message": "Project lon hon 50 bai, dang xu ly nen...",
            }
        comparisons = self.compare_document_against_project(project_id, document_id)
        return {
            "document_id": document_id,
            "title": final_title,
            "queued": False,
            "warnings": warnings,
            "results": comparisons,
        }

    def add_document_from_url(
        self,
        project_id: int,
        url: str,
        *,
        title: str = "",
        source_only: bool = False,
    ) -> dict[str, Any]:
        text, fetch_warnings = fetch_google_doc_text(url)
        result = self.add_document_from_text(project_id, title, text, source_url=url, source_only=source_only)
        result["warnings"] = fetch_warnings + result.get("warnings", [])
        return result

    def list_documents(self, project_id: int) -> list[dict[str, Any]]:
        with self.connection() as conn:
            rows = conn.execute(
                "SELECT * FROM documents WHERE project_id = ? ORDER BY added_at DESC",
                (project_id,),
            ).fetchall()
            payload = []
            for row in rows:
                red_count = conn.execute(
                    """
                    SELECT COUNT(*) AS total
                    FROM comparisons
                    WHERE project_id = ? AND (doc_a_id = ? OR doc_b_id = ?) AND level = 'conflict'
                    """,
                    (project_id, row["id"], row["id"]),
                ).fetchone()["total"]
                yellow_count = conn.execute(
                    """
                    SELECT COUNT(*) AS total
                    FROM comparisons
                    WHERE project_id = ? AND (doc_a_id = ? OR doc_b_id = ?) AND level = 'review'
                    """,
                    (project_id, row["id"], row["id"]),
                ).fetchone()["total"]
                payload.append(
                    {
                        "id": int(row["id"]),
                        "title": row["title"],
                        "added_at": row["added_at"],
                        "source_url": row["source_url"],
                        "doc_role": row["doc_role"] or "check",
                        "approval_status": row["approval_status"] or "approved",
                        "role_label": "Nguồn đối chiếu" if (row["doc_role"] or "check") == "source" else "Bài check",
                        "warning_count": len(json.loads(row["warnings_json"] or "[]")),
                        "conflict_count": int(red_count),
                        "review_count": int(yellow_count),
                        "quality_status": "conflict" if int(red_count) else ("review" if int(yellow_count) else "pass"),
                        "status": "🔴" if int(red_count) else ("🟡" if int(yellow_count) else "✅"),
                    }
                )
            return payload

    def approve_document(self, document_id: int) -> dict[str, Any]:
        with self.connection() as conn:
            row = conn.execute("SELECT id, project_id, title FROM documents WHERE id = ?", (document_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay bai")
            conn.execute(
                """
                UPDATE documents
                SET doc_role = 'source', approval_status = 'approved'
                WHERE id = ?
                """,
                (document_id,),
            )
            conn.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (utc_now(), int(row["project_id"])))
        return {"id": int(row["id"]), "title": row["title"], "status": "approved"}

    def reject_document(self, document_id: int) -> dict[str, Any]:
        with self.connection() as conn:
            row = conn.execute("SELECT id, project_id, title FROM documents WHERE id = ?", (document_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay bai")
            conn.execute(
                """
                UPDATE documents
                SET doc_role = 'check', approval_status = 'rejected'
                WHERE id = ?
                """,
                (document_id,),
            )
            conn.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (utc_now(), int(row["project_id"])))
        return {"id": int(row["id"]), "title": row["title"], "status": "rejected"}

    def remove_document(self, document_id: int) -> None:
        with self.connection() as conn:
            row = conn.execute("SELECT project_id FROM documents WHERE id = ?", (document_id,)).fetchone()
            if not row:
                return
            project_id = int(row["project_id"])
            conn.execute("DELETE FROM documents WHERE id = ?", (document_id,))
            conn.execute("DELETE FROM comparisons WHERE doc_a_id = ? OR doc_b_id = ?", (document_id, document_id))
            conn.execute("DELETE FROM embeddings_cache WHERE document_id = ?", (document_id,))
            conn.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (utc_now(), project_id))

    def recheck_document(self, document_id: int) -> list[dict[str, Any]]:
        with self.connection() as conn:
            row = conn.execute("SELECT project_id FROM documents WHERE id = ?", (document_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay bai")
            project_id = int(row["project_id"])
            conn.execute("DELETE FROM comparisons WHERE doc_a_id = ? OR doc_b_id = ?", (document_id, document_id))
        return self.compare_document_against_project(project_id, document_id)

    def recheck_project(self, project_id: int) -> None:
        with self.connection() as conn:
            conn.execute("DELETE FROM comparisons WHERE project_id = ?", (project_id,))
            docs = conn.execute("SELECT id FROM documents WHERE project_id = ? ORDER BY added_at ASC", (project_id,)).fetchall()
        for row in docs:
            self.compare_document_against_project(project_id, int(row["id"]))

    def _set_job(self, project_id: int, status: str, current: int, total: int, message: str) -> None:
        with self.connection() as conn:
            conn.execute(
                """
                INSERT INTO project_jobs (project_id, status, progress_current, progress_total, message, updated_at)
                VALUES (?, ?, ?, ?, ?, ?)
                ON CONFLICT(project_id) DO UPDATE SET
                    status = excluded.status,
                    progress_current = excluded.progress_current,
                    progress_total = excluded.progress_total,
                    message = excluded.message,
                    updated_at = excluded.updated_at
                """,
                (project_id, status, current, total, message, utc_now()),
            )

    def get_job(self, project_id: int) -> dict[str, Any]:
        with self.connection() as conn:
            row = conn.execute("SELECT * FROM project_jobs WHERE project_id = ?", (project_id,)).fetchone()
        if not row:
            return {"status": "idle", "progress_current": 0, "progress_total": 0, "message": ""}
        return dict(row)

    def _compare_document_background(self, project_id: int, document_id: int) -> None:
        try:
            self.compare_document_against_project(project_id, document_id, background=True)
            self._set_job(project_id, "done", 1, 1, "Da so sanh xong")
        except Exception as exc:  # pragma: no cover - background execution
            self._set_job(project_id, "error", 0, 1, f"Loi: {exc}")

    def _base_product_name(self, title: str, config: dict[str, Any]) -> str:
        working = normalize_compare(title)
        for entry in config.get("allowed_patterns_regex", []):
            regex = entry.get("regex", "") if isinstance(entry, dict) else str(entry)
            if regex:
                working = re.sub(regex, " ", working, flags=re.IGNORECASE)
        working = re.sub(r"\b\d+\b", " ", working)
        tokens = [token for token in tokenize(working) if token not in STOPWORDS]
        return " ".join(tokens[:6])

    def _embedding_for_section(
        self,
        conn: sqlite3.Connection,
        project_id: int,
        document_id: int,
        section: str,
        text: str,
        *,
        multilingual: bool,
    ) -> list[float] | None:
        model_name = "paraphrase-multilingual-MiniLM-L12-v2" if multilingual else "paraphrase-MiniLM-L6-v2"
        digest = text_hash(text)
        cached = conn.execute(
            """
            SELECT embedding_json
            FROM embeddings_cache
            WHERE document_id = ? AND section = ? AND model_name = ? AND text_hash = ?
            """,
            (document_id, section, model_name, digest),
        ).fetchone()
        if cached:
            return json.loads(cached["embedding_json"])
        if not self.semantic_model.available():
            return None
        model = self.semantic_model._load(multilingual)
        if model is None:
            return None
        embedding = model.encode([text], normalize_embeddings=True)[0].tolist()
        conn.execute(
            """
            INSERT INTO embeddings_cache (project_id, document_id, section, model_name, text_hash, embedding_json, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(document_id, section, model_name) DO UPDATE SET
                text_hash = excluded.text_hash,
                embedding_json = excluded.embedding_json,
                created_at = excluded.created_at
            """,
            (project_id, document_id, section, model_name, digest, json.dumps(embedding), utc_now()),
        )
        return embedding

    def _semantic_score(
        self,
        conn: sqlite3.Connection,
        project_id: int,
        doc_a_id: int,
        doc_b_id: int,
        section: str,
        text_a: str,
        text_b: str,
        *,
        multilingual: bool,
    ) -> tuple[float, str]:
        if multilingual and self.semantic_model.available():
            emb_a = self._embedding_for_section(conn, project_id, doc_a_id, section, text_a, multilingual=True)
            emb_b = self._embedding_for_section(conn, project_id, doc_b_id, section, text_b, multilingual=True)
            if emb_a is not None and emb_b is not None:
                score = sum(emb_a[index] * emb_b[index] for index in range(len(emb_a))) * 100.0
                return max(0.0, min(100.0, score)), "semantic_only_multilingual"
        if self.semantic_model.available():
            emb_a = self._embedding_for_section(conn, project_id, doc_a_id, section, text_a, multilingual=False)
            emb_b = self._embedding_for_section(conn, project_id, doc_b_id, section, text_b, multilingual=False)
            if emb_a is not None and emb_b is not None:
                score = sum(emb_a[index] * emb_b[index] for index in range(len(emb_a))) * 100.0
                return max(0.0, min(100.0, score)), "semantic"
        counter_a = Counter(semantic_tokens(text_a))
        counter_b = Counter(semantic_tokens(text_b))
        cosine = cosine_similarity_from_counters(counter_a, counter_b) * 100.0
        sentence_ratio = SequenceMatcher(None, semantic_normalize(text_a), semantic_normalize(text_b)).ratio() * 100.0
        jaccard_tokens = set(counter_a) & set(counter_b)
        union = set(counter_a) | set(counter_b)
        jaccard = (len(jaccard_tokens) / max(1, len(union))) * 100.0
        fallback_score = (cosine * 0.45) + (sentence_ratio * 0.30) + (jaccard * 0.25)
        return fallback_score, "fallback"

    def compare_document_against_project(self, project_id: int, document_id: int, *, background: bool = False) -> list[dict[str, Any]]:
        with self.project_locks[project_id]:
            with self.connection() as conn:
                config = self._project_config(conn, project_id)
                doc_a = conn.execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
                if not doc_a:
                    raise ValueError("Khong tim thay bai")
                sections_a = json.loads(doc_a["sections_json"])
                others = conn.execute(
                    """
                    SELECT * FROM documents
                    WHERE project_id = ? AND id != ? AND doc_role = 'source'
                    ORDER BY added_at ASC
                    """,
                    (project_id, document_id),
                ).fetchall()
                total = len(others)
                results: list[dict[str, Any]] = []
                for index, doc_b in enumerate(others, start=1):
                    if background:
                        self._set_job(project_id, "running", index, total, f"Da check {index}/{total} bai")
                    sections_b = json.loads(doc_b["sections_json"])
                    outcomes = self._compare_documents(conn, project_id, doc_a, sections_a, doc_b, sections_b, config)
                    for outcome in outcomes:
                        conn.execute(
                            """
                            INSERT INTO comparisons (project_id, doc_a_id, doc_b_id, section, risk_score, flagged_sentences_json, created_at, level)
                            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                            """,
                            (
                                project_id,
                                document_id,
                                int(doc_b["id"]),
                                outcome.section,
                                outcome.risk_score,
                                json.dumps(outcome.flagged_sentences, ensure_ascii=False),
                                utc_now(),
                                outcome.level,
                            ),
                        )
                        results.append(
                            {
                                "doc_a_id": document_id,
                                "doc_b_id": int(doc_b["id"]),
                                "doc_b_title": doc_b["title"],
                                "section": outcome.section,
                                "risk_score": round(outcome.risk_score, 2),
                                "level": outcome.level,
                                "detail": outcome.flagged_sentences,
                            }
                        )
                conn.execute("UPDATE projects SET updated_at = ? WHERE id = ?", (utc_now(), project_id))
                return results

    def _compare_documents(
        self,
        conn: sqlite3.Connection,
        project_id: int,
        doc_a: sqlite3.Row,
        sections_a: dict[str, Any],
        doc_b: sqlite3.Row,
        sections_b: dict[str, Any],
        config: dict[str, Any],
    ) -> list[ComparisonOutcome]:
        outcomes: list[ComparisonOutcome] = []
        base_a = self._base_product_name(doc_a["title"], config)
        base_b = self._base_product_name(doc_b["title"], config)
        lenient_variant = bool(base_a and base_b and base_a == base_b and doc_a["title"] != doc_b["title"])
        for section_name, payload_a in sections_a.items():
            canonical = payload_a.get("canonical_name", canonical_section_name(section_name))
            if canonical == "specifications":
                continue
            payload_b = None
            for candidate_name, candidate_payload in sections_b.items():
                if candidate_payload.get("canonical_name", canonical_section_name(candidate_name)) == canonical:
                    payload_b = candidate_payload
                    break
            if not payload_b:
                continue
            outcome = self._compare_section_pair(
                conn,
                project_id,
                doc_a,
                payload_a,
                doc_b,
                payload_b,
                canonical,
                config,
                lenient_variant=lenient_variant,
            )
            if outcome is not None:
                outcomes.append(outcome)
        return outcomes

    def _compare_section_pair(
        self,
        conn: sqlite3.Connection,
        project_id: int,
        doc_a: sqlite3.Row,
        payload_a: dict[str, Any],
        doc_b: sqlite3.Row,
        payload_b: dict[str, Any],
        section_name: str,
        config: dict[str, Any],
        *,
        lenient_variant: bool,
    ) -> ComparisonOutcome | None:
        strip_a = payload_a["strip"]
        strip_b = payload_b["strip"]
        if strip_a["status"] == "allowed" or strip_b["status"] == "allowed":
            return ComparisonOutcome(
                section=section_name,
                risk_score=0.0,
                level="allowed",
                flagged_sentences={
                    "message": "Allowed zone - khong check",
                    "status": "allowed",
                    "strip_log_a": strip_a["strip_log"],
                    "strip_log_b": strip_b["strip_log"],
                },
            )
        if strip_a["status"] == "skip" or strip_b["status"] == "skip":
            return ComparisonOutcome(
                section=section_name,
                risk_score=0.0,
                level="skip",
                flagged_sentences={
                    "message": "Bo qua - qua ngan",
                    "status": "skip",
                    "strip_log_a": strip_a["strip_log"],
                    "strip_log_b": strip_b["strip_log"],
                },
            )
        text_a = strip_a["protected_text"]
        text_b = strip_b["protected_text"]
        tokens_a = meaningful_tokens(text_a)
        tokens_b = meaningful_tokens(text_b)
        if len(tokens_a) < 12 or len(tokens_b) < 12:
            return ComparisonOutcome(
                section=section_name,
                risk_score=0.0,
                level="skip",
                flagged_sentences={"message": "Bo qua - qua ngan", "status": "skip"},
            )
        sentences_a = sent_tokenize_safe(text_a)
        sentences_b = sent_tokenize_safe(text_b)
        first_same = bool(sentences_a and sentences_b and normalize_compare(sentences_a[0]) == normalize_compare(sentences_b[0]))
        last_same = bool(sentences_a and sentences_b and normalize_compare(sentences_a[-1]) == normalize_compare(sentences_b[-1]))
        trigrams_a = word_trigrams(tokens_a)
        trigrams_b = word_trigrams(tokens_b)
        shorter_ngrams = trigrams_a if len(trigrams_a) <= len(trigrams_b) else trigrams_b
        score_ngram = (len(trigrams_a & trigrams_b) / max(1, len(shorter_ngrams))) * 100.0 if shorter_ngrams else 0.0
        lang_a = detect_language_safe(text_a)
        lang_b = detect_language_safe(text_b)
        cross_language = lang_a not in {"unknown", lang_b} and lang_b != "unknown"
        score_semantic, semantic_mode = self._semantic_score(
            conn,
            project_id,
            int(doc_a["id"]),
            int(doc_b["id"]),
            section_name,
            text_a,
            text_b,
            multilingual=cross_language,
        )
        score_lcs, common_sentences = longest_common_sentence_score(sentences_a, sentences_b)
        same_industry_only = False
        shared_terms = {token for token in meaningful_tokens(text_a) if token in meaningful_tokens(text_b)}
        industry_only_terms = {term.replace(" ", "") for term in INDUSTRY_TERMS}
        if shared_terms and all(token in industry_only_terms for token in shared_terms):
            same_industry_only = True
        if cross_language:
            if semantic_mode == "fallback":
                final_score = 0.0
            else:
                final_score = score_semantic
            layer_weights = {"ngram": 0.0, "semantic": 1.0, "lcs": 0.0}
        elif semantic_mode == "fallback" and not self.semantic_model.available():
            final_score = (score_ngram * 0.6) + (score_lcs * 0.4)
            layer_weights = {"ngram": 0.6, "semantic": 0.0, "lcs": 0.4}
        else:
            final_score = (score_ngram * 0.4) + (score_semantic * 0.4) + (score_lcs * 0.2)
            layer_weights = {"ngram": 0.4, "semantic": 0.4, "lcs": 0.2}
        if min(len(tokens_a), len(tokens_b)) < 30:
            final_score *= min(len(tokens_a), len(tokens_b)) / 30.0
        threshold = float(config.get("section_thresholds", {}).get(section_name, DEFAULT_THRESHOLDS["unknown_section"]))
        if lenient_variant:
            threshold *= 1.1
        note_list: list[str] = []
        if lenient_variant:
            note_list.append("Day co the la san pham cung dong, trung lap template la binh thuong")
        if same_industry_only and score_ngram < 35 and score_semantic < 75:
            final_score = 0.0
        if score_semantic > 60.0 and score_ngram < 20.0:
            if score_lcs > 30.0:
                final_score = max(final_score, threshold * 0.8)
                note_list.append("Paraphrase ro rang nhung van co cau trung lap")
            else:
                final_score = min(final_score, threshold * 0.5)
                note_list.append("Paraphrase tot - semantic cao nhung wording khac")
        level = "ok"
        if first_same:
            level = "conflict"
            final_score = max(final_score, threshold + 5)
            note_list.append("Cung cau dau tien")
        elif last_same:
            level = "review"
            final_score = max(final_score, threshold * 0.8)
            note_list.append("Cung cau cuoi")
        if round(final_score, 2) >= 100.0:
            note_list.append("Co the la cung bai duoc paste lai")
        if round(final_score, 2) == 0.0 and score_semantic > 60.0:
            note_list.append("Paraphrase ro rang")
        if level == "ok":
            if final_score >= threshold:
                level = "conflict"
            elif final_score >= threshold * 0.75:
                level = "review"
        preview_left, preview_right = best_matching_sentence(sentences_a, sentences_b)
        return ComparisonOutcome(
            section=section_name,
            risk_score=max(0.0, min(100.0, final_score)),
            level=level,
            flagged_sentences={
                "status": level,
                "preview_a": preview_left,
                "preview_b": preview_right,
                "layer_scores": {
                    "ngram": round(score_ngram, 2),
                    "semantic": round(score_semantic, 2),
                    "lcs": round(score_lcs, 2),
                    "weights": layer_weights,
                    "semantic_mode": semantic_mode,
                },
                "common_sentences": common_sentences[:3],
                "strip_log_a": strip_a["strip_log"],
                "strip_log_b": strip_b["strip_log"],
                "protected_a": text_a,
                "protected_b": text_b,
                "raw_a": payload_a["text"],
                "raw_b": payload_b["text"],
                "threshold": round(threshold, 2),
                "notes": note_list,
            },
        )

    def project_results(
        self,
        project_id: int,
        *,
        doc_id: int | None = None,
        level: str = "",
        section: str = "",
        other_doc_id: int | None = None,
        search: str = "",
        sort_by: str = "risk",
    ) -> list[dict[str, Any]]:
        with self.connection() as conn:
            rows = conn.execute(
                """
                SELECT c.*, da.title AS doc_a_title, db.title AS doc_b_title
                FROM comparisons c
                JOIN documents da ON da.id = c.doc_a_id
                JOIN documents db ON db.id = c.doc_b_id
                WHERE c.project_id = ?
                """,
                (project_id,),
            ).fetchall()
        results = []
        for row in rows:
            detail = json.loads(row["flagged_sentences_json"] or "{}")
            if doc_id and int(row["doc_a_id"]) != doc_id and int(row["doc_b_id"]) != doc_id:
                continue
            if level and row["level"] != level:
                continue
            if section and row["section"] != section:
                continue
            if other_doc_id and int(row["doc_a_id"]) != other_doc_id and int(row["doc_b_id"]) != other_doc_id:
                continue
            searchable = " ".join(
                [
                    row["doc_a_title"],
                    row["doc_b_title"],
                    row["section"],
                    detail.get("preview_a", ""),
                    detail.get("preview_b", ""),
                ]
            ).lower()
            if search and search.lower() not in searchable:
                continue
            results.append(
                {
                    "id": int(row["id"]),
                    "doc_a_id": int(row["doc_a_id"]),
                    "doc_b_id": int(row["doc_b_id"]),
                    "doc_a_title": row["doc_a_title"],
                    "doc_b_title": row["doc_b_title"],
                    "section": row["section"],
                    "risk_score": round(float(row["risk_score"]), 2),
                    "level": row["level"],
                    "detail": detail,
                    "created_at": row["created_at"],
                }
            )
        if sort_by == "section":
            results.sort(key=lambda item: (item["section"], -item["risk_score"]))
        elif sort_by == "doc":
            results.sort(key=lambda item: (item["doc_b_title"].lower(), -item["risk_score"]))
        else:
            results.sort(key=lambda item: item["risk_score"], reverse=True)
        return results

    def export_excel(self, project_id: int) -> bytes:
        if openpyxl is None or PatternFill is None:
            raise RuntimeError("Can cai openpyxl de xuat Excel")
        project = self.get_project(project_id)
        docs = self.list_documents(project_id)
        results = self.project_results(project_id)
        wb = openpyxl.Workbook()
        ws1 = wb.active
        ws1.title = "Tong quan conflict"
        ws1.append(["STT", "Bai moi", "Bai xung dot", "Section", "Risk Score", "Muc do", "Cau trung tieu bieu", "Ngay check"])
        fills = {
            "conflict": PatternFill("solid", fgColor="FFCCCC"),
            "review": PatternFill("solid", fgColor="FFFACC"),
            "ok": PatternFill("solid", fgColor="CCFFCC"),
            "allowed": PatternFill("solid", fgColor="CCFFCC"),
            "skip": PatternFill("solid", fgColor="CCFFCC"),
        }
        for index, result in enumerate(results, start=1):
            preview = result["detail"].get("preview_a") or result["detail"].get("message", "")
            ws1.append(
                [
                    index,
                    result["doc_a_title"],
                    result["doc_b_title"],
                    result["section"],
                    result["risk_score"],
                    RESULT_LEVELS.get(result["level"], result["level"]),
                    preview,
                    result["created_at"],
                ]
            )
            for cell in ws1[ws1.max_row]:
                cell.fill = fills.get(result["level"], fills["ok"])

        ws2 = wb.create_sheet("Danh sach tat ca bai")
        ws2.append(["Bai", "Ngay them", "So conflict 🔴", "So xem lai 🟡", "Tinh trang tong"])
        for doc in docs:
            ws2.append([doc["title"], doc["added_at"], doc["conflict_count"], doc["review_count"], doc["status"]])
            status_key = "conflict" if doc["status"] == "🔴" else ("review" if doc["status"] == "🟡" else "ok")
            for cell in ws2[ws2.max_row]:
                cell.fill = fills.get(status_key, fills["ok"])

        ws3 = wb.create_sheet("Allowed Zone da config")
        ws3.append(["Pattern", "Loai", "Vi du tim thay trong template"])
        config = project["allowed_zone_config"]
        for entry in config.get("allowed_patterns_regex", []):
            ws3.append([entry.get("regex", ""), "regex", ", ".join(entry.get("examples", []))])
        for phrase in config.get("allowed_phrases", []):
            ws3.append([phrase, "phrase", phrase])
        for section_name in config.get("fully_allowed_sections", []):
            ws3.append([section_name, "section", section_name])

        output = io.BytesIO()
        wb.save(output)
        return output.getvalue()

    def export_project_zip(self, project_id: int) -> bytes:
        project = self.get_project(project_id)
        docs = self.list_documents(project_id)
        results = self.project_results(project_id)
        with self.connection() as conn:
            project_row = conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()
            document_rows = conn.execute(
                "SELECT * FROM documents WHERE project_id = ? ORDER BY added_at ASC",
                (project_id,),
            ).fetchall()
            comparison_rows = conn.execute(
                "SELECT * FROM comparisons WHERE project_id = ? ORDER BY created_at ASC, id ASC",
                (project_id,),
            ).fetchall()
        output = io.BytesIO()
        with zipfile.ZipFile(output, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("project.json", json.dumps(project, ensure_ascii=False, indent=2))
            archive.writestr("documents.json", json.dumps(docs, ensure_ascii=False, indent=2))
            archive.writestr("comparisons.json", json.dumps(results, ensure_ascii=False, indent=2))
            archive.writestr("allowed_zone.json", json.dumps(project["allowed_zone_config"], ensure_ascii=False, indent=2))
            archive.writestr(
                "project_full.json",
                json.dumps(dict(project_row) if project_row else {}, ensure_ascii=False, indent=2),
            )
            archive.writestr(
                "documents_full.json",
                json.dumps([dict(row) for row in document_rows], ensure_ascii=False, indent=2),
            )
            archive.writestr(
                "comparisons_full.json",
                json.dumps([dict(row) for row in comparison_rows], ensure_ascii=False, indent=2),
            )
        return output.getvalue()

    def import_project_zip(self, payload: bytes) -> dict[str, Any]:
        try:
            with zipfile.ZipFile(io.BytesIO(payload), "r") as archive:
                project_full = json.loads(archive.read("project_full.json").decode("utf-8"))
                documents_full = json.loads(archive.read("documents_full.json").decode("utf-8"))
                comparisons_full = json.loads(archive.read("comparisons_full.json").decode("utf-8"))
        except KeyError as exc:
            raise ValueError("Backup ZIP thiếu dữ liệu cần thiết để khôi phục") from exc
        except zipfile.BadZipFile as exc:
            raise ValueError("File ZIP không hợp lệ") from exc

        new_name = normalize_space(project_full.get("name") or "Project khôi phục")
        new_project = self.create_project(f"{new_name} (khôi phục)")
        id_map: dict[int, int] = {}
        with self.connection() as conn:
            conn.execute(
                """
                UPDATE projects
                SET template_doc_text = ?, allowed_zone_config_json = ?, updated_at = ?
                WHERE id = ?
                """,
                (
                    project_full.get("template_doc_text", ""),
                    project_full.get("allowed_zone_config_json", "{}"),
                    utc_now(),
                    new_project["id"],
                ),
            )
            for row in documents_full:
                cursor = conn.execute(
                    """
                    INSERT INTO documents (
                        project_id, title, raw_text, sections_json, added_at,
                        content_hash, source_url, doc_role, approval_status, warnings_json
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_project["id"],
                        row.get("title", "Tài liệu"),
                        row.get("raw_text", ""),
                        row.get("sections_json", "{}"),
                        row.get("added_at", utc_now()),
                        row.get("content_hash", text_hash(row.get("raw_text", ""))),
                        row.get("source_url", ""),
                        row.get("doc_role", "check"),
                        row.get("approval_status", "approved"),
                        row.get("warnings_json", "[]"),
                    ),
                )
                old_id = int(row.get("id", 0))
                if old_id:
                    id_map[old_id] = int(cursor.lastrowid)
            for row in comparisons_full:
                old_doc_a = int(row.get("doc_a_id", 0))
                old_doc_b = int(row.get("doc_b_id", 0))
                if old_doc_a not in id_map or old_doc_b not in id_map:
                    continue
                conn.execute(
                    """
                    INSERT INTO comparisons (
                        project_id, doc_a_id, doc_b_id, section, risk_score, flagged_sentences_json, created_at, level
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        new_project["id"],
                        id_map[old_doc_a],
                        id_map[old_doc_b],
                        row.get("section", "unknown_section"),
                        float(row.get("risk_score", 0.0)),
                        row.get("flagged_sentences_json", "{}"),
                        row.get("created_at", utc_now()),
                        row.get("level", "ok"),
                    ),
                )
        return self.get_project(new_project["id"])

    def self_test(self) -> dict[str, Any]:
        temp_handle = tempfile.NamedTemporaryFile(prefix="self_test_", suffix=".sqlite3", dir=DATA_DIR, delete=False)
        temp_handle.close()
        temp_db = Path(temp_handle.name)
        try:
            service = SmartDuplicateService(temp_db)
            project = service.create_project("Test")
            template_text = """
            PRODUCT TITLE

            INTRO
            This product is designed for professional takeaway and delivery use. It supports fast-moving kitchens with a dependable packaging format.

            FEATURES
            The container stays stackable and supports secure lid closure during delivery service. The material also helps keep portions tidy during packing.

            SPECIFICATIONS
            900ml
            168mm
            300 units

            USE CASES
            Ideal for salad bowls, noodle bowls, and rice meal service in busy lunch and dinner operations.
            """
            service.set_template_from_text(project["id"], template_text)
            doc_base = service.add_document_from_text(
                project["id"],
                "Doc base",
                """
                Product Title

                Intro
                This product is designed for professional takeaway and delivery use. It supports fast-moving kitchens with a dependable packaging format.

                Features
                The container stays stackable and supports secure lid closure during delivery service. The material also helps keep portions tidy during packing.

                Specifications
                900ml
                168mm
                300 units

                Use Cases
                Ideal for salad bowls, noodle bowls, and rice meal service in busy lunch and dinner operations.
                """,
                source_only=True,
            )
            doc_identical = service.add_document_from_text(
                project["id"],
                "Doc identical",
                """
                Product Title 900ml

                Intro
                This product is designed for professional takeaway and delivery use in busy kitchens. It supports fast-moving kitchens with a dependable packaging format.

                Features
                The container stays stackable and supports secure lid closure during delivery service. The material also helps keep portions tidy during packing.

                Specifications
                1090ml
                168mm
                300 units

                Use Cases
                Ideal for salad bowls, noodle bowls, and rice meal service in busy lunch and dinner operations.
                """,
            )
            doc_paraphrase = service.add_document_from_text(
                project["id"],
                "Doc paraphrase",
                """
                Product Title Variant

                Intro
                Built for takeaway operations, this item supports professional delivery workflows and suits high-volume kitchen packing lines.

                Features
                The container remains stackable and supports secure lid closure during delivery service. The material also helps keep portions tidy during packing.

                Specifications
                1200ml
                168mm
                300 units

                Use Cases
                Suitable for salad bowls, noodle portions, and rice meal service across busy lunch and dinner operations.
                """,
            )
            doc_different = service.add_document_from_text(
                project["id"],
                "Doc different",
                """
                Another Product

                Intro
                A clear deli pot intended for cold desserts and fruit display cabinets in retail environments.

                Features
                The transparent body improves visibility on shelf and suits chilled retail presentation for front-of-house merchandising.

                Specifications
                12oz
                95mm
                500 units

                Use Cases
                Best for parfaits, fruit cups, and dessert toppings rather than hot takeaway meal service.
                """,
            )
            results = service.project_results(project["id"])
            return {
                "project_id": project["id"],
                "docs_added": [
                    doc_base["document_id"],
                    doc_identical["document_id"],
                    doc_paraphrase["document_id"],
                    doc_different["document_id"],
                ],
                "results_total": len(results),
                "identical_expected": any(
                    item["doc_a_title"] == "Doc identical" and item["doc_b_title"] == "Doc base" and item["level"] == "conflict"
                    for item in results
                ),
                "paraphrase_expected": any(
                    item["doc_a_title"] == "Doc paraphrase"
                    and item["doc_b_title"] in {"Doc base", "Doc identical"}
                    and item["level"] in {"review", "conflict"}
                    for item in results
                ),
                "different_expected": any(
                    item["doc_a_title"] == "Doc different" and item["level"] == "ok"
                    for item in results
                ),
                "spec_guard_expected": not any(
                    item["section"] == "specifications" and item["level"] in {"conflict", "review"} for item in results
                ),
                "results": results,
                "note": "Neu sentence-transformers chua co, semantic layer dang fallback.",
            }
        finally:
            if temp_db.exists():
                temp_db.unlink(missing_ok=True)
