"""Storage, import, and scheduling logic for the SEO Brain app."""

from __future__ import annotations

import hashlib
import io
import json
import math
import re
import sqlite3
from collections import defaultdict
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


DB_FILE_NAME = "seo_brain.sqlite3"
DEFAULT_TEAM_CAPACITY = {
    "SEO": 30.0,
    "Content": 45.0,
    "Editor": 22.0,
    "Client": 10.0,
    "Outreach": 18.0,
    "Dev": 14.0,
}
DEFAULT_MONTH_COUNT = 10
PRIORITY_SCORES = {"P1": 3, "P2": 2, "P3": 1}
STATUS_TO_STAGE = {
    "Published": "Publish",
    "To review (Client)": "Client Review",
    "To review (Internal)": "Internal Review",
    "Doing": "Draft",
    "Processing": "Draft",
    "To do": "Keyword Map",
    "Not started": "Keyword Map",
}
METRIC_DEFINITIONS = {
    "publish_pages": {"label": "Pages live", "unit": "pages", "goal": "higher"},
    "content_ready": {"label": "Content ready", "unit": "pages", "goal": "higher"},
    "onpage_ready": {"label": "Onpage ready", "unit": "pages", "goal": "higher"},
    "indexed_pages": {"label": "Indexed pages", "unit": "pages", "goal": "higher"},
    "index_rate": {"label": "Index rate", "unit": "%", "goal": "higher"},
    "offpage_live": {"label": "Offpage support", "unit": "pages", "goal": "higher"},
    "critical_blockers": {"label": "Critical blockers", "unit": "pages", "goal": "lower"},
}
WORKFLOW_OFFSETS: dict[str, list[dict[str, Any]]] = {
    "product": [
        {"task_name": "Keyword Map", "workstream": "Research", "role": "SEO", "effort_hours": 1.5, "start_offset": -14, "due_offset": -12, "success_check": "Primary keyword, intent, and target URL approved."},
        {"task_name": "Outline", "workstream": "Content", "role": "Content", "effort_hours": 2.0, "start_offset": -11, "due_offset": -9, "success_check": "Outline approved with headings, USP, FAQ, and internal-link opportunities."},
        {"task_name": "Draft", "workstream": "Content", "role": "Content", "effort_hours": 4.0, "start_offset": -8, "due_offset": -5, "success_check": "Draft complete with entity coverage, media brief, and proof points."},
        {"task_name": "Internal Review", "workstream": "Content", "role": "Editor", "effort_hours": 1.5, "start_offset": -4, "due_offset": -3, "success_check": "Internal QA passes structure, tone, and SEO requirements."},
        {"task_name": "Client Review", "workstream": "Content", "role": "Client", "effort_hours": 1.0, "start_offset": -2, "due_offset": -1, "success_check": "Client feedback resolved and approved for upload."},
        {"task_name": "Upload CMS", "workstream": "Content", "role": "Content", "effort_hours": 1.0, "start_offset": -1, "due_offset": -1, "success_check": "Draft uploaded with correct modules, media, and formatting."},
        {"task_name": "Onpage SEO", "workstream": "Onpage", "role": "SEO", "effort_hours": 2.0, "start_offset": 0, "due_offset": 0, "success_check": "H1, title, meta, schema, canonical, images, and CWV notes checked."},
        {"task_name": "Publish", "workstream": "Launch", "role": "SEO", "effort_hours": 0.5, "start_offset": 0, "due_offset": 0, "success_check": "URL is live and renders correctly on production."},
        {"task_name": "Request Indexing", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.3, "start_offset": 1, "due_offset": 1, "success_check": "URL submitted in GSC or indexation queue updated."},
        {"task_name": "Index Check D+3", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.4, "start_offset": 3, "due_offset": 3, "success_check": "Index status checked on day 3 and re-submit flagged if needed."},
        {"task_name": "Internal Links", "workstream": "Internal Link", "role": "SEO", "effort_hours": 1.5, "start_offset": 2, "due_offset": 5, "success_check": "Supportive internal links placed from relevant pages and tracked."},
        {"task_name": "Offpage Launch", "workstream": "Offpage", "role": "Outreach", "effort_hours": 2.5, "start_offset": 7, "due_offset": 14, "success_check": "Offpage support asset or placement plan is live for this URL."},
        {"task_name": "Impact Review", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.8, "start_offset": 21, "due_offset": 28, "success_check": "Index, rank, CTR, and crawl signals reviewed with next action logged."},
    ],
    "category": [
        {"task_name": "Keyword Map", "workstream": "Research", "role": "SEO", "effort_hours": 2.0, "start_offset": -18, "due_offset": -16, "success_check": "Primary and supporting clusters mapped to the page."},
        {"task_name": "Outline", "workstream": "Content", "role": "Content", "effort_hours": 3.0, "start_offset": -15, "due_offset": -12, "success_check": "Section plan covers commercial intent, filters, FAQs, and trust blocks."},
        {"task_name": "Draft", "workstream": "Content", "role": "Content", "effort_hours": 6.0, "start_offset": -11, "due_offset": -7, "success_check": "Draft complete with category copy, semantic clusters, and media asks."},
        {"task_name": "Internal Review", "workstream": "Content", "role": "Editor", "effort_hours": 2.0, "start_offset": -6, "due_offset": -5, "success_check": "Internal QA passes structure, commercial angle, and differentiation."},
        {"task_name": "Client Review", "workstream": "Content", "role": "Client", "effort_hours": 1.5, "start_offset": -4, "due_offset": -3, "success_check": "Client signs off or requested edits are resolved."},
        {"task_name": "Upload CMS", "workstream": "Content", "role": "Content", "effort_hours": 1.0, "start_offset": -2, "due_offset": -2, "success_check": "Page content uploaded and modules/UX blocks are in place."},
        {"task_name": "Onpage SEO", "workstream": "Onpage", "role": "SEO", "effort_hours": 3.0, "start_offset": -1, "due_offset": 0, "success_check": "Titles, schema, indexability, filters, and copy placement are checked."},
        {"task_name": "Publish", "workstream": "Launch", "role": "SEO", "effort_hours": 0.5, "start_offset": 0, "due_offset": 0, "success_check": "Category page is live on production."},
        {"task_name": "Request Indexing", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.3, "start_offset": 1, "due_offset": 1, "success_check": "Index request sent and sitemap updated if needed."},
        {"task_name": "Index Check D+3", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.4, "start_offset": 3, "due_offset": 3, "success_check": "Day-3 index signal confirmed or escalated."},
        {"task_name": "Internal Links", "workstream": "Internal Link", "role": "SEO", "effort_hours": 2.0, "start_offset": 2, "due_offset": 6, "success_check": "Cluster pages point to this category page with tracked anchors."},
        {"task_name": "Offpage Launch", "workstream": "Offpage", "role": "Outreach", "effort_hours": 3.0, "start_offset": 7, "due_offset": 16, "success_check": "Supportive outreach/backlink plan launched for the category page."},
        {"task_name": "Impact Review", "workstream": "Measurement", "role": "SEO", "effort_hours": 1.0, "start_offset": 21, "due_offset": 30, "success_check": "Rank/click/index movement reviewed and next sprint task logged."},
    ],
    "blog": [
        {"task_name": "Keyword Map", "workstream": "Research", "role": "SEO", "effort_hours": 2.0, "start_offset": -21, "due_offset": -18, "success_check": "Primary query, intent, SERP angle, and conversion path approved."},
        {"task_name": "Outline", "workstream": "Content", "role": "Content", "effort_hours": 3.0, "start_offset": -17, "due_offset": -14, "success_check": "Outline includes entity coverage, supporting questions, and link targets."},
        {"task_name": "Draft", "workstream": "Content", "role": "Content", "effort_hours": 8.0, "start_offset": -13, "due_offset": -8, "success_check": "Draft complete with examples, FAQs, and internal-link hooks."},
        {"task_name": "Internal Review", "workstream": "Content", "role": "Editor", "effort_hours": 2.5, "start_offset": -7, "due_offset": -6, "success_check": "Editorial + SEO QA completed with revisions resolved."},
        {"task_name": "Client Review", "workstream": "Content", "role": "Client", "effort_hours": 1.5, "start_offset": -5, "due_offset": -4, "success_check": "Client comments cleared or waived."},
        {"task_name": "Upload CMS", "workstream": "Content", "role": "Content", "effort_hours": 1.0, "start_offset": -3, "due_offset": -3, "success_check": "Article uploaded with media, CTA, and formatting clean."},
        {"task_name": "Onpage SEO", "workstream": "Onpage", "role": "SEO", "effort_hours": 2.5, "start_offset": -2, "due_offset": 0, "success_check": "SEO title, meta, headings, schema, and image SEO pass."},
        {"task_name": "Publish", "workstream": "Launch", "role": "SEO", "effort_hours": 0.5, "start_offset": 0, "due_offset": 0, "success_check": "Blog URL is live and visible in sitemap/feed."},
        {"task_name": "Request Indexing", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.3, "start_offset": 1, "due_offset": 1, "success_check": "Index request logged for the article."},
        {"task_name": "Index Check D+3", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.4, "start_offset": 3, "due_offset": 3, "success_check": "Index status checked after 3 days and escalated if needed."},
        {"task_name": "Internal Links", "workstream": "Internal Link", "role": "SEO", "effort_hours": 1.5, "start_offset": 2, "due_offset": 5, "success_check": "Article links to and from priority pages with tracked anchors."},
        {"task_name": "Offpage Launch", "workstream": "Offpage", "role": "Outreach", "effort_hours": 2.0, "start_offset": 7, "due_offset": 14, "success_check": "Supporting distribution or outreach launched for the article."},
        {"task_name": "Impact Review", "workstream": "Measurement", "role": "SEO", "effort_hours": 0.8, "start_offset": 21, "due_offset": 28, "success_check": "Index and early traffic/rank signals reviewed."},
    ],
}


def now_iso() -> str:
    return datetime.utcnow().replace(microsecond=0).isoformat()


def safe_slug(value: str) -> str:
    slug = re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-")
    return slug or "item"


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def parse_float(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        text = normalize_text(value).replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None


def parse_excel_date(value: Any, default_year: int) -> date | None:
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = normalize_text(value)
    if not text or text in {"-", "—"}:
        return None
    text = text.split("00:00:00")[0].strip()
    if " - " in text and "/" in text:
        text = text.split(" - ", 1)[1].strip()
    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return datetime.strptime(text, "%Y-%m-%d").date()
    if re.fullmatch(r"\d{1,2}/\d{1,2}/\d{4}", text):
        return datetime.strptime(text, "%d/%m/%Y").date()
    if re.fullmatch(r"\d{1,2}/\d{1,2}", text):
        day_value, month_value = text.split("/")
        return date(default_year, int(month_value), int(day_value))
    return None


def format_date(value: str | None) -> str:
    if not value:
        return ""
    try:
        parsed = date.fromisoformat(value)
    except ValueError:
        return value
    return parsed.strftime("%d/%m/%Y")


def add_workdays(start: date, offset: int) -> date:
    current = start
    direction = 1 if offset >= 0 else -1
    remaining = abs(offset)
    while remaining:
        current += timedelta(days=direction)
        if current.weekday() < 5:
            remaining -= 1
    return current


def workday_distance(start: date, end: date) -> int:
    if start == end:
        return 0
    direction = 1 if end > start else -1
    current = start
    distance = 0
    while current != end:
        current += timedelta(days=direction)
        if current.weekday() < 5:
            distance += direction
    return distance


def end_of_month(month_start: date) -> date:
    if month_start.month == 12:
        return date(month_start.year + 1, 1, 1) - timedelta(days=1)
    return date(month_start.year, month_start.month + 1, 1) - timedelta(days=1)


def month_key_for(value: date) -> str:
    return value.strftime("%Y-%m")


def workflow_kind(page_type: str) -> str:
    lower = page_type.lower()
    if "blog" in lower or "case" in lower:
        return "blog"
    if "category" in lower or "catalog" in lower or "cms" in lower:
        return "category"
    return "product"


class SeoBrainService:
    def __init__(self, base_dir: Path):
        self.base_dir = base_dir
        data_dir = base_dir / "data"
        data_dir.mkdir(parents=True, exist_ok=True)
        self.db_path = data_dir / DB_FILE_NAME
        self._ensure_schema()

    def _connect(self) -> sqlite3.Connection:
        connection = sqlite3.connect(self.db_path)
        connection.row_factory = sqlite3.Row
        connection.execute("PRAGMA foreign_keys = ON")
        return connection

    def _ensure_schema(self) -> None:
        with self._connect() as conn:
            conn.executescript(
                """
                CREATE TABLE IF NOT EXISTS projects (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    name TEXT NOT NULL,
                    source_name TEXT NOT NULL,
                    source_hash TEXT NOT NULL,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    last_imported_at TEXT NOT NULL
                );
                CREATE TABLE IF NOT EXISTS project_settings (
                    project_id INTEGER PRIMARY KEY,
                    team_capacity_json TEXT NOT NULL,
                    kickoff_date TEXT,
                    notes TEXT DEFAULT '',
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE TABLE IF NOT EXISTS pages (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    external_key TEXT,
                    category TEXT,
                    page_type TEXT,
                    priority TEXT,
                    title TEXT,
                    primary_keyword TEXT,
                    search_volume REAL,
                    url TEXT,
                    content_status TEXT,
                    draft_window TEXT,
                    review_deadline TEXT,
                    publish_target TEXT,
                    index_status TEXT,
                    pic TEXT,
                    notes TEXT,
                    image_status TEXT,
                    image_priority REAL,
                    stock_status TEXT,
                    health_score REAL DEFAULT 0,
                    row_order INTEGER DEFAULT 0,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_pages_project ON pages(project_id);
                CREATE TABLE IF NOT EXISTS tasks (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    page_id INTEGER,
                    task_key TEXT NOT NULL,
                    workstream TEXT NOT NULL,
                    task_name TEXT NOT NULL,
                    role_name TEXT NOT NULL,
                    owner TEXT,
                    sequence_no INTEGER NOT NULL,
                    planned_start TEXT,
                    planned_due TEXT,
                    forecast_start TEXT,
                    forecast_due TEXT,
                    actual_due TEXT,
                    actual_done TEXT,
                    status TEXT NOT NULL,
                    effort_hours REAL NOT NULL DEFAULT 0,
                    buffer_days INTEGER NOT NULL DEFAULT 0,
                    dependency_keys TEXT NOT NULL DEFAULT '[]',
                    warning_level TEXT NOT NULL DEFAULT 'none',
                    impact_score REAL NOT NULL DEFAULT 0,
                    success_check TEXT,
                    notes TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE,
                    FOREIGN KEY(page_id) REFERENCES pages(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_tasks_project ON tasks(project_id);
                CREATE INDEX IF NOT EXISTS idx_tasks_page ON tasks(page_id);
                CREATE TABLE IF NOT EXISTS kpis (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    project_id INTEGER NOT NULL,
                    month_key TEXT NOT NULL,
                    metric_key TEXT NOT NULL,
                    metric_name TEXT NOT NULL,
                    unit TEXT NOT NULL,
                    target_value REAL,
                    actual_override REAL,
                    linked_workstream TEXT,
                    notes TEXT,
                    created_at TEXT NOT NULL,
                    updated_at TEXT NOT NULL,
                    UNIQUE(project_id, month_key, metric_key),
                    FOREIGN KEY(project_id) REFERENCES projects(id) ON DELETE CASCADE
                );
                CREATE INDEX IF NOT EXISTS idx_kpis_project_month ON kpis(project_id, month_key);
                """
            )

    def list_projects(self) -> list[sqlite3.Row]:
        with self._connect() as conn:
            return conn.execute(
                """
                SELECT
                    p.*,
                    COUNT(DISTINCT pg.id) AS page_count,
                    COUNT(DISTINCT t.id) AS task_count
                FROM projects p
                LEFT JOIN pages pg ON pg.project_id = p.id
                LEFT JOIN tasks t ON t.project_id = p.id
                GROUP BY p.id
                ORDER BY p.updated_at DESC, p.id DESC
                """
            ).fetchall()

    def get_project(self, project_id: int) -> sqlite3.Row | None:
        with self._connect() as conn:
            return conn.execute("SELECT * FROM projects WHERE id = ?", (project_id,)).fetchone()

    def get_project_settings(self, project_id: int) -> dict[str, Any]:
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM project_settings WHERE project_id = ?", (project_id,)).fetchone()
        if not row:
            return {"team_capacity": dict(DEFAULT_TEAM_CAPACITY), "kickoff_date": ""}
        capacities = json.loads(row["team_capacity_json"] or "{}")
        return {
            "team_capacity": {**DEFAULT_TEAM_CAPACITY, **capacities},
            "kickoff_date": row["kickoff_date"] or "",
            "notes": row["notes"] or "",
        }

    def import_workbook(self, source_name: str, workbook_bytes: bytes, project_name: str = "") -> dict[str, Any]:
        workbook = load_workbook(io.BytesIO(workbook_bytes), data_only=True)
        project_title = project_name.strip() or Path(source_name).stem
        imported_at = now_iso()
        source_hash = hashlib.sha1(workbook_bytes).hexdigest()
        content_rows = self._extract_content_rows(workbook)
        if not content_rows:
            raise ValueError("Khong tim thay du lieu o sheet Content Calendar.")

        relevant_dates = [
            row["publish_target_obj"]
            for row in content_rows
            if row["publish_target_obj"] is not None
        ]
        relevant_dates.extend(
            row["review_deadline_obj"]
            for row in content_rows
            if row["review_deadline_obj"] is not None
        )
        kickoff = min(relevant_dates) if relevant_dates else date.today()

        with self._connect() as conn:
            cursor = conn.execute(
                """
                INSERT INTO projects (name, source_name, source_hash, created_at, updated_at, last_imported_at)
                VALUES (?, ?, ?, ?, ?, ?)
                """,
                (project_title, source_name, source_hash, imported_at, imported_at, imported_at),
            )
            project_id = int(cursor.lastrowid)
            conn.execute(
                """
                INSERT INTO project_settings (project_id, team_capacity_json, kickoff_date, notes)
                VALUES (?, ?, ?, '')
                """,
                (project_id, json.dumps(DEFAULT_TEAM_CAPACITY), kickoff.isoformat()),
            )
            image_map = self._extract_image_audit_map(workbook, kickoff.year)
            pages: list[dict[str, Any]] = []
            for row in content_rows:
                payload = {
                    "project_id": project_id,
                    "external_key": normalize_text(row["external_key"]),
                    "category": row["category"],
                    "page_type": row["page_type"],
                    "priority": row["priority"] or "P3",
                    "title": row["title"],
                    "primary_keyword": row["primary_keyword"],
                    "search_volume": row["search_volume"],
                    "url": row["url"],
                    "content_status": row["content_status"] or "To do",
                    "draft_window": row["draft_window"],
                    "review_deadline": row["review_deadline"],
                    "publish_target": row["publish_target"],
                    "index_status": row["index_status"] or "Not checked",
                    "pic": row["pic"],
                    "notes": row["notes"],
                    "image_status": "",
                    "image_priority": None,
                    "stock_status": "",
                    "health_score": 0,
                    "row_order": row["row_order"],
                    "created_at": imported_at,
                    "updated_at": imported_at,
                }
                if not payload["publish_target"]:
                    payload["publish_target"] = self._page_anchor_date(payload, kickoff).isoformat()
                image_row = image_map.get((row["url"] or "").strip())
                if image_row:
                    payload["image_status"] = image_row["image_status"]
                    payload["image_priority"] = image_row["priority"]
                    payload["stock_status"] = image_row["stock_status"]
                cursor = conn.execute(
                    """
                    INSERT INTO pages (
                        project_id, external_key, category, page_type, priority, title,
                        primary_keyword, search_volume, url, content_status, draft_window,
                        review_deadline, publish_target, index_status, pic, notes,
                        image_status, image_priority, stock_status, health_score, row_order,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        payload["project_id"],
                        payload["external_key"],
                        payload["category"],
                        payload["page_type"],
                        payload["priority"],
                        payload["title"],
                        payload["primary_keyword"],
                        payload["search_volume"],
                        payload["url"],
                        payload["content_status"],
                        payload["draft_window"],
                        payload["review_deadline"],
                        payload["publish_target"],
                        payload["index_status"],
                        payload["pic"],
                        payload["notes"],
                        payload["image_status"],
                        payload["image_priority"],
                        payload["stock_status"],
                        payload["health_score"],
                        payload["row_order"],
                        payload["created_at"],
                        payload["updated_at"],
                    ),
                )
                payload["id"] = int(cursor.lastrowid)
                pages.append(payload)

            tasks = []
            for page_row in pages:
                tasks.extend(self._build_tasks_for_page(page_row, kickoff))

            for task in tasks:
                conn.execute(
                    """
                    INSERT INTO tasks (
                        project_id, page_id, task_key, workstream, task_name, role_name, owner,
                        sequence_no, planned_start, planned_due, forecast_start, forecast_due,
                        actual_due, actual_done, status, effort_hours, buffer_days,
                        dependency_keys, warning_level, impact_score, success_check, notes,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        task["project_id"],
                        task["page_id"],
                        task["task_key"],
                        task["workstream"],
                        task["task_name"],
                        task["role_name"],
                        task["owner"],
                        task["sequence_no"],
                        task["planned_start"],
                        task["planned_due"],
                        task["forecast_start"],
                        task["forecast_due"],
                        task["actual_due"],
                        task["actual_done"],
                        task["status"],
                        task["effort_hours"],
                        task["buffer_days"],
                        json.dumps(task["dependency_keys"]),
                        task["warning_level"],
                        task["impact_score"],
                        task["success_check"],
                        task["notes"],
                        imported_at,
                        imported_at,
                    ),
                )
            conn.commit()

        self.recompute_project(project_id)
        self.seed_default_kpis(project_id)
        summary = self.get_project_summary(project_id)
        return {
            "project_id": project_id,
            "project_name": project_title,
            "pages_imported": summary["total_pages"],
            "tasks_generated": summary["total_tasks"],
        }

    def _extract_content_rows(self, workbook: Any) -> list[dict[str, Any]]:
        if "Content Calendar" not in workbook.sheetnames:
            return []
        ws = workbook["Content Calendar"]
        rows: list[dict[str, Any]] = []
        default_year = self._guess_workbook_year(workbook)
        for row_index, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
            url = normalize_text(row[11] if len(row) > 11 else "")
            title = normalize_text(row[4] if len(row) > 4 else "")
            primary_keyword = normalize_text(row[8] if len(row) > 8 else "")
            page_type = normalize_text(row[2] if len(row) > 2 else "")
            if not any((url, title, primary_keyword, page_type)):
                continue
            publish_target_obj = parse_excel_date(row[15] if len(row) > 15 else None, default_year)
            review_deadline_obj = parse_excel_date(row[14] if len(row) > 14 else None, default_year)
            derived_title = title or primary_keyword or Path(url.strip("/")).name.replace("-", " ").title() or f"Row {row_index}"
            rows.append(
                {
                    "row_order": row_index,
                    "external_key": row[0] if len(row) > 0 else row_index,
                    "category": normalize_text(row[1] if len(row) > 1 else ""),
                    "page_type": page_type or "Page",
                    "priority": normalize_text(row[3] if len(row) > 3 else "") or "P3",
                    "title": derived_title,
                    "primary_keyword": primary_keyword,
                    "search_volume": parse_float(row[9] if len(row) > 9 else None),
                    "url": url.strip(),
                    "content_status": normalize_text(row[12] if len(row) > 12 else "") or "To do",
                    "draft_window": normalize_text(row[13] if len(row) > 13 else ""),
                    "review_deadline": review_deadline_obj.isoformat() if review_deadline_obj else "",
                    "publish_target": publish_target_obj.isoformat() if publish_target_obj else (review_deadline_obj.isoformat() if review_deadline_obj else ""),
                    "review_deadline_obj": review_deadline_obj,
                    "publish_target_obj": publish_target_obj or review_deadline_obj,
                    "index_status": normalize_text(row[17] if len(row) > 17 else ""),
                    "notes": normalize_text(row[18] if len(row) > 18 else ""),
                    "pic": normalize_text(row[19] if len(row) > 19 else ""),
                }
            )
        return rows

    def _extract_image_audit_map(self, workbook: Any, default_year: int) -> dict[str, dict[str, Any]]:
        if "SEO IMAGE AUDIT" not in workbook.sheetnames:
            return {}
        ws = workbook["SEO IMAGE AUDIT"]
        result: dict[str, dict[str, Any]] = {}
        for row in ws.iter_rows(min_row=4, values_only=True):
            url = normalize_text(row[3] if len(row) > 3 else "")
            if not url:
                continue
            result[url.strip()] = {
                "priority": parse_float(row[4] if len(row) > 4 else None),
                "stock_status": normalize_text(row[5] if len(row) > 5 else ""),
                "image_status": normalize_text(row[6] if len(row) > 6 else "") or normalize_text(row[19] if len(row) > 19 else ""),
            }
        return result

    def _guess_workbook_year(self, workbook: Any) -> int:
        for sheet_name in ("Content Calendar", "Timeline", "Bi-weekly Report"):
            if sheet_name not in workbook.sheetnames:
                continue
            ws = workbook[sheet_name]
            for row in ws.iter_rows(min_row=1, max_row=15, values_only=True):
                for value in row:
                    if isinstance(value, datetime):
                        return value.year
                    if isinstance(value, date):
                        return value.year
        return date.today().year

    def _build_tasks_for_page(self, page_row: dict[str, Any], kickoff: date) -> list[dict[str, Any]]:
        kind = workflow_kind(page_row["page_type"])
        workflow = WORKFLOW_OFFSETS[kind]
        publish_anchor = self._page_anchor_date(page_row, kickoff)
        current_stage_name = STATUS_TO_STAGE.get(page_row["content_status"], "")
        current_stage_order = next(
            (index for index, item in enumerate(workflow) if item["task_name"] == current_stage_name),
            -1,
        )
        tasks: list[dict[str, Any]] = []
        previous_task_key = ""
        publish_task_key = ""
        index_check_task_key = ""
        offpage_task_key = ""
        for sequence_no, template in enumerate(workflow, start=1):
            planned_start = add_workdays(publish_anchor, int(template["start_offset"]))
            planned_due = add_workdays(publish_anchor, int(template["due_offset"]))
            task_key = safe_slug(
                f"p{page_row['id']}-{template['task_name']}-{page_row['title']}-{page_row['priority']}"
            )
            dependency_keys: list[str] = []
            if previous_task_key:
                dependency_keys = [previous_task_key]
            if template["task_name"] in {"Internal Links", "Offpage Launch"} and publish_task_key:
                dependency_keys = [publish_task_key]
            if template["task_name"] == "Impact Review":
                dependency_keys = [item for item in (index_check_task_key, offpage_task_key or publish_task_key) if item]
            status = "todo"
            actual_done = ""
            if sequence_no - 1 <= current_stage_order:
                status = "done"
                actual_done = planned_due.isoformat()
            if template["task_name"] == "Index Check D+3" and "index" in (page_row["index_status"] or "").lower():
                status = "done"
                actual_done = planned_due.isoformat()
            task = {
                "project_id": page_row["project_id"],
                "page_id": page_row["id"],
                "task_key": task_key,
                "workstream": template["workstream"],
                "task_name": template["task_name"],
                "role_name": template["role"],
                "owner": page_row["pic"] or template["role"],
                "sequence_no": sequence_no,
                "planned_start": planned_start.isoformat(),
                "planned_due": planned_due.isoformat(),
                "forecast_start": planned_start.isoformat(),
                "forecast_due": planned_due.isoformat(),
                "actual_due": "",
                "actual_done": actual_done,
                "status": status,
                "effort_hours": template["effort_hours"],
                "buffer_days": 0,
                "dependency_keys": list(dependency_keys),
                "warning_level": "done" if status == "done" else "none",
                "impact_score": 0.0,
                "success_check": template["success_check"],
                "notes": "",
            }
            tasks.append(task)
            previous_task_key = task_key
            if template["task_name"] == "Publish":
                publish_task_key = task_key
            if template["task_name"] == "Index Check D+3":
                index_check_task_key = task_key
            if template["task_name"] == "Offpage Launch":
                offpage_task_key = task_key
        return tasks

    def _page_anchor_date(self, page_row: dict[str, Any], kickoff: date) -> date:
        if page_row.get("publish_target"):
            return date.fromisoformat(page_row["publish_target"])
        if page_row.get("review_deadline"):
            return add_workdays(date.fromisoformat(page_row["review_deadline"]), 2)
        priority_offset = {"P1": 10, "P2": 25, "P3": 45}.get(page_row.get("priority") or "P3", 45)
        return add_workdays(kickoff, priority_offset + int(page_row.get("row_order", 0) % 10))

    def seed_default_kpis(self, project_id: int) -> None:
        months = self.project_months(project_id)
        derived = self.derived_monthly_metrics(project_id)
        timestamp = now_iso()
        with self._connect() as conn:
            for month in months:
                month_key = month["month_key"]
                metrics = derived.get(month_key, {})
                for metric_key, meta in METRIC_DEFINITIONS.items():
                    values = metrics.get(metric_key, {})
                    target_value = values.get("target")
                    conn.execute(
                        """
                        INSERT INTO kpis (
                            project_id, month_key, metric_key, metric_name, unit,
                            target_value, actual_override, linked_workstream, notes,
                            created_at, updated_at
                        )
                        VALUES (?, ?, ?, ?, ?, ?, NULL, ?, '', ?, ?)
                        ON CONFLICT(project_id, month_key, metric_key) DO UPDATE SET
                            metric_name = excluded.metric_name,
                            unit = excluded.unit,
                            target_value = COALESCE(kpis.target_value, excluded.target_value),
                            linked_workstream = excluded.linked_workstream,
                            updated_at = excluded.updated_at
                        """,
                        (
                            project_id,
                            month_key,
                            metric_key,
                            meta["label"],
                            meta["unit"],
                            target_value,
                            self._default_workstream_for_metric(metric_key),
                            timestamp,
                            timestamp,
                        ),
                    )
            conn.commit()

    def _default_workstream_for_metric(self, metric_key: str) -> str:
        mapping = {
            "publish_pages": "Launch",
            "content_ready": "Content",
            "onpage_ready": "Onpage",
            "indexed_pages": "Measurement",
            "index_rate": "Measurement",
            "offpage_live": "Offpage",
            "critical_blockers": "All",
        }
        return mapping.get(metric_key, "All")

    def project_months(self, project_id: int) -> list[dict[str, Any]]:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT MIN(COALESCE(publish_target, review_deadline)) AS start_value,
                       MAX(COALESCE(publish_target, review_deadline)) AS end_value
                FROM pages WHERE project_id = ?
                """,
                (project_id,),
            ).fetchone()
        if row and row["start_value"]:
            start_date = date.fromisoformat(row["start_value"])
        else:
            settings = self.get_project_settings(project_id)
            start_date = date.fromisoformat(settings.get("kickoff_date") or date.today().isoformat())
        if row and row["end_value"]:
            end_date = date.fromisoformat(row["end_value"])
        else:
            end_date = add_workdays(start_date, 90)
        months: list[dict[str, Any]] = []
        cursor = date(start_date.year, start_date.month, 1)
        max_months = 0
        while cursor <= date(end_date.year, end_date.month, 1) or max_months < DEFAULT_MONTH_COUNT:
            months.append(
                {
                    "month_key": cursor.strftime("%Y-%m"),
                    "label": cursor.strftime("%b %Y"),
                    "start": cursor.isoformat(),
                    "end": end_of_month(cursor).isoformat(),
                }
            )
            max_months += 1
            if cursor.month == 12:
                cursor = date(cursor.year + 1, 1, 1)
            else:
                cursor = date(cursor.year, cursor.month + 1, 1)
            if max_months >= 18:
                break
        return months

    def derived_monthly_metrics(self, project_id: int) -> dict[str, dict[str, dict[str, float | str]]]:
        months = self.project_months(project_id)
        month_map = {item["month_key"]: item for item in months}
        with self._connect() as conn:
            task_rows = conn.execute(
                """
                SELECT t.*, p.priority, p.index_status, p.publish_target, p.content_status
                FROM tasks t
                LEFT JOIN pages p ON p.id = t.page_id
                WHERE t.project_id = ?
                ORDER BY t.planned_due, t.sequence_no
                """,
                (project_id,),
            ).fetchall()
        data = {
            month["month_key"]: {
                metric_key: {"target": 0.0, "actual": 0.0, "forecast": 0.0}
                for metric_key in METRIC_DEFINITIONS
            }
            for month in months
        }
        publish_actual_by_month: defaultdict[str, float] = defaultdict(float)
        index_actual_by_month: defaultdict[str, float] = defaultdict(float)
        blockers_actual: defaultdict[str, set[int]] = defaultdict(set)
        blockers_forecast: defaultdict[str, set[int]] = defaultdict(set)
        for row in task_rows:
            planned_due = date.fromisoformat(row["planned_due"]) if row["planned_due"] else None
            forecast_due = date.fromisoformat(row["forecast_due"]) if row["forecast_due"] else planned_due
            actual_done = date.fromisoformat(row["actual_done"]) if row["actual_done"] else None
            if row["task_name"] == "Publish":
                month_key = month_key_for(planned_due or forecast_due or date.today())
                if month_key in data:
                    data[month_key]["publish_pages"]["target"] += 1
                    data[month_key]["onpage_ready"]["target"] += 1
                    data[month_key]["content_ready"]["target"] += 1
                    data[month_key]["indexed_pages"]["target"] += 1
                    data[month_key]["offpage_live"]["target"] += 1
                forecast_month = month_key_for(forecast_due or planned_due or date.today())
                if forecast_month in data:
                    data[forecast_month]["publish_pages"]["forecast"] += 1
                if actual_done:
                    actual_month = month_key_for(actual_done)
                    publish_actual_by_month[actual_month] += 1
                    if actual_month in data:
                        data[actual_month]["publish_pages"]["actual"] += 1
            if row["task_name"] == "Client Review":
                month_key = month_key_for(planned_due or forecast_due or date.today())
                if month_key in data:
                    data[month_key]["content_ready"]["forecast"] += 1
                    if row["status"] == "done":
                        data[month_key]["content_ready"]["actual"] += 1
            if row["task_name"] == "Onpage SEO":
                month_key = month_key_for(forecast_due or planned_due or date.today())
                if month_key in data:
                    data[month_key]["onpage_ready"]["forecast"] += 1
                    if row["status"] == "done":
                        done_month = month_key_for(actual_done or forecast_due or planned_due or date.today())
                        if done_month in data:
                            data[done_month]["onpage_ready"]["actual"] += 1
            if row["task_name"] == "Index Check D+3":
                month_key = month_key_for(forecast_due or planned_due or date.today())
                if month_key in data:
                    data[month_key]["indexed_pages"]["forecast"] += 1
                if row["status"] == "done":
                    actual_month = month_key_for(actual_done or forecast_due or planned_due or date.today())
                    index_actual_by_month[actual_month] += 1
                    if actual_month in data:
                        data[actual_month]["indexed_pages"]["actual"] += 1
            if row["task_name"] == "Offpage Launch":
                month_key = month_key_for(forecast_due or planned_due or date.today())
                if month_key in data:
                    data[month_key]["offpage_live"]["forecast"] += 1
                    if row["status"] == "done":
                        done_month = month_key_for(actual_done or forecast_due or planned_due or date.today())
                        if done_month in data:
                            data[done_month]["offpage_live"]["actual"] += 1
            if (
                row["warning_level"] == "red"
                and row["task_name"] in {"Draft", "Client Review", "Onpage SEO", "Publish", "Index Check D+3"}
                and row["priority"] in {"P1", "P2"}
                and row["page_id"]
            ):
                month_key = month_key_for(forecast_due or planned_due or date.today())
                blockers_forecast[month_key].add(int(row["page_id"]))
                if row["status"] == "done":
                    blockers_actual[month_key].add(int(row["page_id"]))

        cumulative_published = 0.0
        cumulative_indexed = 0.0
        for month in months:
            month_key = month["month_key"]
            cumulative_published += publish_actual_by_month.get(month_key, 0.0)
            cumulative_indexed += index_actual_by_month.get(month_key, 0.0)
            data[month_key]["index_rate"]["target"] = 80.0
            data[month_key]["index_rate"]["actual"] = (
                round((cumulative_indexed / cumulative_published) * 100, 1) if cumulative_published else 0.0
            )
            forecast_published = sum(
                data[other["month_key"]]["publish_pages"]["forecast"]
                for other in months
                if other["month_key"] <= month_key
            )
            forecast_indexed = sum(
                data[other["month_key"]]["indexed_pages"]["forecast"]
                for other in months
                if other["month_key"] <= month_key
            )
            data[month_key]["index_rate"]["forecast"] = (
                round((forecast_indexed / forecast_published) * 100, 1) if forecast_published else 0.0
            )
            data[month_key]["critical_blockers"]["actual"] = float(len(blockers_actual.get(month_key, set())))
            data[month_key]["critical_blockers"]["forecast"] = float(len(blockers_forecast.get(month_key, set())))
        return data

    def recompute_project(self, project_id: int) -> None:
        settings = self.get_project_settings(project_id)
        capacities = {**DEFAULT_TEAM_CAPACITY, **settings.get("team_capacity", {})}
        kickoff_raw = settings.get("kickoff_date") or date.today().isoformat()
        kickoff = date.fromisoformat(kickoff_raw)
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT t.*, p.priority, p.search_volume, p.url, p.index_status, p.image_status, p.publish_target
                FROM tasks t
                LEFT JOIN pages p ON p.id = t.page_id
                WHERE t.project_id = ?
                ORDER BY p.priority = 'P1' DESC, p.priority = 'P2' DESC, p.search_volume DESC, t.planned_due, t.sequence_no
                """,
                (project_id,),
            ).fetchall()
            tasks = [dict(row) for row in rows]

        role_loads: dict[str, dict[date, float]] = defaultdict(dict)
        task_lookup: dict[str, dict[str, Any]] = {}
        updated_tasks: list[dict[str, Any]] = []
        for task in tasks:
            dependencies = json.loads(task["dependency_keys"] or "[]")
            planned_start = date.fromisoformat(task["planned_start"]) if task["planned_start"] else kickoff
            planned_due = date.fromisoformat(task["planned_due"]) if task["planned_due"] else planned_start
            actual_done = date.fromisoformat(task["actual_done"]) if task["actual_done"] else None
            actual_due = date.fromisoformat(task["actual_due"]) if task["actual_due"] else None
            dep_dates = []
            for dep_key in dependencies:
                dep_task = task_lookup.get(dep_key)
                if dep_task and dep_task.get("forecast_due"):
                    dep_dates.append(date.fromisoformat(dep_task["forecast_due"]))
            earliest = planned_start
            if dep_dates:
                earliest = max(earliest, add_workdays(max(dep_dates), 1))
            role = task["role_name"]
            daily_capacity = max(capacities.get(role, 8.0) / 5, 0.5)

            if task["status"] == "done":
                forecast_start = planned_start
                completion = actual_done or actual_due or planned_due
                forecast_due = completion
                warning_level = "done"
            else:
                forecast_start, forecast_due = self._allocate_role_capacity(
                    role_loads=role_loads,
                    role=role,
                    earliest=earliest,
                    effort_hours=float(task["effort_hours"] or 0),
                    daily_capacity=daily_capacity,
                )
                slip_days = workday_distance(planned_due, forecast_due)
                warning_level = "green"
                if slip_days > 2:
                    warning_level = "yellow"
                if slip_days > 7:
                    warning_level = "orange"
                if slip_days > 14:
                    warning_level = "red"
                if any(task_lookup.get(dep_key, {}).get("warning_level") == "red" for dep_key in dependencies):
                    warning_level = "red"

            priority_score = PRIORITY_SCORES.get(task.get("priority") or "P3", 1)
            dependent_weight = 1 + len(dependencies)
            slip_value = max(0, workday_distance(planned_due, forecast_due))
            impact_score = round(priority_score * dependent_weight * (1 + slip_value / 2), 1)
            task["forecast_start"] = forecast_start.isoformat()
            task["forecast_due"] = forecast_due.isoformat()
            task["warning_level"] = warning_level
            task["impact_score"] = impact_score
            updated_tasks.append(task)
            task_lookup[task["task_key"]] = task

        with self._connect() as conn:
            for task in updated_tasks:
                conn.execute(
                    """
                    UPDATE tasks
                    SET forecast_start = ?, forecast_due = ?, warning_level = ?, impact_score = ?, updated_at = ?
                    WHERE id = ?
                    """,
                    (
                        task["forecast_start"],
                        task["forecast_due"],
                        task["warning_level"],
                        task["impact_score"],
                        now_iso(),
                        task["id"],
                    ),
                )
            conn.commit()
        self._recompute_page_health(project_id)

    def _allocate_role_capacity(
        self,
        *,
        role_loads: dict[str, dict[date, float]],
        role: str,
        earliest: date,
        effort_hours: float,
        daily_capacity: float,
    ) -> tuple[date, date]:
        remaining_effort = max(effort_hours, 0.25)
        current = earliest
        start_day: date | None = None
        role_calendar = role_loads[role]
        while True:
            if current.weekday() >= 5:
                current += timedelta(days=1)
                continue
            available = role_calendar.get(current, daily_capacity)
            if available <= 0:
                current += timedelta(days=1)
                continue
            if start_day is None:
                start_day = current
            work_chunk = min(available, remaining_effort)
            available -= work_chunk
            remaining_effort -= work_chunk
            role_calendar[current] = available
            if remaining_effort <= 0.001:
                return start_day, current
            current += timedelta(days=1)

    def _recompute_page_health(self, project_id: int) -> None:
        with self._connect() as conn:
            pages = conn.execute("SELECT * FROM pages WHERE project_id = ?", (project_id,)).fetchall()
            for page_row in pages:
                tasks = conn.execute(
                    "SELECT * FROM tasks WHERE project_id = ? AND page_id = ? ORDER BY sequence_no",
                    (project_id, page_row["id"]),
                ).fetchall()
                score = 100.0
                if any(task["warning_level"] == "red" for task in tasks):
                    score -= 30
                if any(task["warning_level"] == "orange" for task in tasks):
                    score -= 15
                if page_row["content_status"] not in {"Published", "Done"}:
                    score -= 10
                if "index" not in (page_row["index_status"] or "").lower():
                    score -= 10
                image_status = (page_row["image_status"] or "").lower()
                if any(marker in image_status for marker in ("thieu", "missing", "chua")):
                    score -= 10
                if (page_row["stock_status"] or "").lower().startswith("out of stock"):
                    score -= 5
                score = max(20.0, min(100.0, score))
                conn.execute(
                    "UPDATE pages SET health_score = ?, updated_at = ? WHERE id = ?",
                    (score, now_iso(), page_row["id"]),
                )
            conn.commit()

    def get_project_summary(self, project_id: int) -> dict[str, Any]:
        with self._connect() as conn:
            page_row = conn.execute(
                """
                SELECT
                    COUNT(*) AS total_pages,
                    SUM(CASE WHEN priority = 'P1' THEN 1 ELSE 0 END) AS p1_pages,
                    SUM(CASE WHEN content_status = 'Published' THEN 1 ELSE 0 END) AS published_pages,
                    AVG(health_score) AS average_health
                FROM pages WHERE project_id = ?
                """,
                (project_id,),
            ).fetchone()
            task_row = conn.execute(
                """
                SELECT
                    COUNT(*) AS total_tasks,
                    SUM(CASE WHEN status = 'done' THEN 1 ELSE 0 END) AS completed_tasks,
                    SUM(CASE WHEN warning_level = 'red' THEN 1 ELSE 0 END) AS red_tasks,
                    SUM(CASE WHEN warning_level = 'orange' THEN 1 ELSE 0 END) AS orange_tasks
                FROM tasks WHERE project_id = ?
                """,
                (project_id,),
            ).fetchone()
        return {
            "total_pages": int(page_row["total_pages"] or 0),
            "p1_pages": int(page_row["p1_pages"] or 0),
            "published_pages": int(page_row["published_pages"] or 0),
            "average_health": round(float(page_row["average_health"] or 0), 1),
            "total_tasks": int(task_row["total_tasks"] or 0),
            "completed_tasks": int(task_row["completed_tasks"] or 0),
            "red_tasks": int(task_row["red_tasks"] or 0),
            "orange_tasks": int(task_row["orange_tasks"] or 0),
        }

    def list_pages(
        self,
        project_id: int,
        *,
        priority: str = "",
        status: str = "",
        query: str = "",
    ) -> list[sqlite3.Row]:
        clauses = ["project_id = ?"]
        params: list[Any] = [project_id]
        if priority:
            clauses.append("priority = ?")
            params.append(priority)
        if status:
            clauses.append("content_status = ?")
            params.append(status)
        if query:
            clauses.append("(title LIKE ? OR primary_keyword LIKE ? OR url LIKE ?)")
            needle = f"%{query}%"
            params.extend([needle, needle, needle])
        with self._connect() as conn:
            return conn.execute(
                f"""
                SELECT * FROM pages
                WHERE {' AND '.join(clauses)}
                ORDER BY priority = 'P1' DESC, priority = 'P2' DESC, health_score ASC, row_order
                """,
                params,
            ).fetchall()

    def list_tasks(
        self,
        project_id: int,
        *,
        month_key: str = "",
        workstream: str = "",
        warning: str = "",
        page_id: int | None = None,
        week_key: str = "",
    ) -> list[sqlite3.Row]:
        clauses = ["t.project_id = ?"]
        params: list[Any] = [project_id]
        if month_key:
            clauses.append("substr(COALESCE(t.forecast_due, t.planned_due), 1, 7) = ?")
            params.append(month_key)
        if workstream:
            clauses.append("t.workstream = ?")
            params.append(workstream)
        if warning:
            clauses.append("t.warning_level = ?")
            params.append(warning)
        if page_id:
            clauses.append("t.page_id = ?")
            params.append(page_id)
        if week_key:
            week_start = date.fromisoformat(week_key)
            week_end = week_start + timedelta(days=6)
            clauses.append("COALESCE(t.forecast_due, t.planned_due) BETWEEN ? AND ?")
            params.extend([week_start.isoformat(), week_end.isoformat()])
        with self._connect() as conn:
            return conn.execute(
                f"""
                SELECT
                    t.*,
                    p.title AS page_title,
                    p.priority,
                    p.url
                FROM tasks t
                LEFT JOIN pages p ON p.id = t.page_id
                WHERE {' AND '.join(clauses)}
                ORDER BY COALESCE(t.forecast_due, t.planned_due), t.warning_level DESC, t.sequence_no
                """,
                params,
            ).fetchall()

    def update_task_field(self, task_id: int, field: str, value: str) -> dict[str, Any]:
        allowed = {"owner", "status", "actual_due", "actual_done", "notes"}
        if field not in allowed:
            raise ValueError("Field khong hop le.")
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM tasks WHERE id = ?", (task_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay task.")
            if field in {"actual_due", "actual_done"} and value:
                value = date.fromisoformat(value).isoformat()
            conn.execute(
                f"UPDATE tasks SET {field} = ?, updated_at = ? WHERE id = ?",
                (value, now_iso(), task_id),
            )
            conn.commit()
            project_id = int(row["project_id"])
        self.recompute_project(project_id)
        return self.task_detail(task_id)

    def update_page_field(self, page_id: int, field: str, value: str) -> dict[str, Any]:
        allowed = {"priority", "content_status", "publish_target", "review_deadline", "index_status", "pic", "notes"}
        if field not in allowed:
            raise ValueError("Field khong hop le.")
        with self._connect() as conn:
            row = conn.execute("SELECT * FROM pages WHERE id = ?", (page_id,)).fetchone()
            if not row:
                raise ValueError("Khong tim thay page.")
            if field in {"publish_target", "review_deadline"} and value:
                value = date.fromisoformat(value).isoformat()
            conn.execute(
                f"UPDATE pages SET {field} = ?, updated_at = ? WHERE id = ?",
                (value, now_iso(), page_id),
            )
            conn.commit()
            project_id = int(row["project_id"])
        if field in {"publish_target", "review_deadline", "priority"}:
            self.rebuild_page_tasks(page_id)
        else:
            self.recompute_project(project_id)
        return self.page_detail(page_id)

    def rebuild_page_tasks(self, page_id: int) -> None:
        with self._connect() as conn:
            page = conn.execute("SELECT * FROM pages WHERE id = ?", (page_id,)).fetchone()
            if not page:
                raise ValueError("Khong tim thay page.")
            existing_tasks = conn.execute(
                "SELECT * FROM tasks WHERE page_id = ? ORDER BY sequence_no",
                (page_id,),
            ).fetchall()
            project_id = int(page["project_id"])
            settings = self.get_project_settings(project_id)
            kickoff = date.fromisoformat(settings.get("kickoff_date") or date.today().isoformat())
            regenerated = self._build_tasks_for_page(dict(page), kickoff)
            by_name = {row["task_name"]: row for row in existing_tasks}
            conn.execute("DELETE FROM tasks WHERE page_id = ?", (page_id,))
            for task in regenerated:
                previous = by_name.get(task["task_name"])
                if previous:
                    task["status"] = previous["status"]
                    task["actual_due"] = previous["actual_due"] or ""
                    task["actual_done"] = previous["actual_done"] or ""
                    task["notes"] = previous["notes"] or ""
                    task["owner"] = previous["owner"] or task["owner"]
                conn.execute(
                    """
                    INSERT INTO tasks (
                        project_id, page_id, task_key, workstream, task_name, role_name, owner,
                        sequence_no, planned_start, planned_due, forecast_start, forecast_due,
                        actual_due, actual_done, status, effort_hours, buffer_days,
                        dependency_keys, warning_level, impact_score, success_check, notes,
                        created_at, updated_at
                    )
                    VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    (
                        task["project_id"],
                        task["page_id"],
                        task["task_key"],
                        task["workstream"],
                        task["task_name"],
                        task["role_name"],
                        task["owner"],
                        task["sequence_no"],
                        task["planned_start"],
                        task["planned_due"],
                        task["forecast_start"],
                        task["forecast_due"],
                        task["actual_due"],
                        task["actual_done"],
                        task["status"],
                        task["effort_hours"],
                        task["buffer_days"],
                        json.dumps(task["dependency_keys"]),
                        task["warning_level"],
                        task["impact_score"],
                        task["success_check"],
                        task["notes"],
                        now_iso(),
                        now_iso(),
                    ),
                )
            conn.commit()
        self.recompute_project(project_id)

    def page_detail(self, page_id: int) -> dict[str, Any]:
        with self._connect() as conn:
            page = conn.execute("SELECT * FROM pages WHERE id = ?", (page_id,)).fetchone()
            if not page:
                raise ValueError("Khong tim thay page.")
            tasks = conn.execute(
                """
                SELECT * FROM tasks
                WHERE page_id = ?
                ORDER BY sequence_no
                """,
                (page_id,),
            ).fetchall()
        publish_task = next((task for task in tasks if task["task_name"] == "Publish"), None)
        return {
            "id": page["id"],
            "title": page["title"],
            "priority": page["priority"],
            "url": page["url"],
            "health_score": page["health_score"],
            "content_status": page["content_status"],
            "publish_target": page["publish_target"] or "",
            "publish_forecast": publish_task["forecast_due"] if publish_task else "",
            "index_status": page["index_status"] or "",
            "image_status": page["image_status"] or "",
            "pic": page["pic"] or "",
            "notes": page["notes"] or "",
        }

    def task_detail(self, task_id: int) -> dict[str, Any]:
        with self._connect() as conn:
            row = conn.execute(
                """
                SELECT t.*, p.title AS page_title
                FROM tasks t
                LEFT JOIN pages p ON p.id = t.page_id
                WHERE t.id = ?
                """,
                (task_id,),
            ).fetchone()
        if not row:
            raise ValueError("Khong tim thay task.")
        return {
            "id": row["id"],
            "task_name": row["task_name"],
            "page_title": row["page_title"] or "",
            "status": row["status"],
            "owner": row["owner"] or "",
            "planned_due": row["planned_due"] or "",
            "forecast_due": row["forecast_due"] or "",
            "actual_due": row["actual_due"] or "",
            "actual_done": row["actual_done"] or "",
            "warning_level": row["warning_level"],
            "impact_score": row["impact_score"],
        }

    def update_team_capacity(self, project_id: int, payload: dict[str, float], kickoff_date: str = "") -> None:
        merged = {**DEFAULT_TEAM_CAPACITY, **payload}
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO project_settings (project_id, team_capacity_json, kickoff_date, notes)
                VALUES (?, ?, ?, '')
                ON CONFLICT(project_id) DO UPDATE SET
                    team_capacity_json = excluded.team_capacity_json,
                    kickoff_date = CASE WHEN excluded.kickoff_date = '' THEN project_settings.kickoff_date ELSE excluded.kickoff_date END
                """,
                (project_id, json.dumps(merged), kickoff_date),
            )
            conn.commit()
        self.recompute_project(project_id)

    def upsert_kpi(
        self,
        project_id: int,
        month_key: str,
        metric_key: str,
        metric_name: str,
        unit: str,
        target_value: float | None,
        actual_override: float | None = None,
        linked_workstream: str = "",
        notes: str = "",
    ) -> None:
        with self._connect() as conn:
            conn.execute(
                """
                INSERT INTO kpis (
                    project_id, month_key, metric_key, metric_name, unit,
                    target_value, actual_override, linked_workstream, notes,
                    created_at, updated_at
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(project_id, month_key, metric_key) DO UPDATE SET
                    metric_name = excluded.metric_name,
                    unit = excluded.unit,
                    target_value = excluded.target_value,
                    actual_override = excluded.actual_override,
                    linked_workstream = excluded.linked_workstream,
                    notes = excluded.notes,
                    updated_at = excluded.updated_at
                """,
                (
                    project_id,
                    month_key,
                    metric_key,
                    metric_name,
                    unit,
                    target_value,
                    actual_override,
                    linked_workstream,
                    notes,
                    now_iso(),
                    now_iso(),
                ),
            )
            conn.commit()

    def monthly_kpi_board(self, project_id: int) -> list[dict[str, Any]]:
        derived = self.derived_monthly_metrics(project_id)
        with self._connect() as conn:
            rows = conn.execute(
                """
                SELECT * FROM kpis
                WHERE project_id = ?
                ORDER BY month_key, metric_name
                """,
                (project_id,),
            ).fetchall()
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            metric_values = derived.get(row["month_key"], {}).get(row["metric_key"], {})
            actual_value = row["actual_override"]
            if actual_value is None:
                actual_value = metric_values.get("actual")
            forecast_value = metric_values.get("forecast")
            target_value = row["target_value"]
            goal = METRIC_DEFINITIONS.get(row["metric_key"], {}).get("goal", "higher")
            status = self._metric_status(goal, target_value, actual_value, forecast_value)
            grouped[row["month_key"]].append(
                {
                    "id": row["id"],
                    "metric_key": row["metric_key"],
                    "metric_name": row["metric_name"],
                    "unit": row["unit"],
                    "target_value": target_value,
                    "actual_value": actual_value,
                    "forecast_value": forecast_value,
                    "linked_workstream": row["linked_workstream"] or "",
                    "notes": row["notes"] or "",
                    "status": status,
                }
            )
        board = []
        for month in self.project_months(project_id):
            board.append(
                {
                    "month_key": month["month_key"],
                    "label": month["label"],
                    "metrics": grouped.get(month["month_key"], []),
                }
            )
        return board

    def _metric_status(
        self,
        goal: str,
        target_value: float | None,
        actual_value: float | None,
        forecast_value: float | None,
    ) -> str:
        if target_value is None:
            return "neutral"
        if goal == "lower":
            if forecast_value is not None:
                if forecast_value <= target_value:
                    if actual_value is not None and actual_value > target_value:
                        return "yellow"
                    return "green"
                return "red"
            if actual_value is not None:
                return "green" if actual_value <= target_value else "red"
            return "neutral"
        if actual_value is not None and actual_value >= target_value:
            return "green"
        if forecast_value is not None and forecast_value >= target_value:
            return "yellow"
        return "red"

    def weekly_plan(self, project_id: int, week_start: date | None = None) -> dict[str, list[dict[str, Any]]]:
        if week_start is None:
            today = date.today()
            week_start = today - timedelta(days=today.weekday())
        rows = self.list_tasks(project_id, week_key=week_start.isoformat())
        grouped: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for row in rows:
            owner = row["owner"] or row["role_name"]
            grouped[owner].append(
                {
                    "id": row["id"],
                    "task_name": row["task_name"],
                    "page_title": row["page_title"] or "",
                    "planned_due": row["planned_due"] or "",
                    "forecast_due": row["forecast_due"] or "",
                    "status": row["status"],
                    "warning_level": row["warning_level"],
                }
            )
        return grouped
